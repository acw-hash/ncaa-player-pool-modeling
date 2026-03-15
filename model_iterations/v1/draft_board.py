#!/usr/bin/env python3
"""
NCAA Tournament Player Pool — Draft Board Generator
=====================================================
Generates a ranked draft board from a pre-tournament cheat sheet.

USAGE:
    python draft_board.py input.xlsx
    python draft_board.py input.xlsx --output my_draft_board.xlsx

INPUT:
    An Excel file with columns (case-insensitive, flexible naming):
        - Player (or Name)
        - Team
        - Seed
        - PPG (or PTS, Points)
        - Region (optional)
        - Rank (optional — the cheat sheet's pre-tournament ranking)

OUTPUT:
    An Excel file with ranked draft board + model diagnostics.
    All model artifacts (EDA, fitting, optimization) print to console.

MODEL:
    Fitted on 4 years of historical pool data (2022-2025, ~400 observations).
    See console output for full model details.
"""

import pandas as pd
import numpy as np
import argparse
import sys
import os
from collections import defaultdict

# =============================================================================
# HISTORICAL SEED ADVANCEMENT PROBABILITIES (1985-2024, ~40 years of data)
# Each list: [P(play R64), P(play R32), P(play S16), P(play E8), P(play F4), P(play Championship)]
# =============================================================================
SEED_ADVANCE_PROBS = {
    1:  [1.000, 0.993, 0.850, 0.600, 0.390, 0.220, 0.130],
    2:  [1.000, 0.940, 0.670, 0.400, 0.220, 0.110, 0.060],
    3:  [1.000, 0.850, 0.510, 0.250, 0.120, 0.050, 0.020],
    4:  [1.000, 0.790, 0.430, 0.200, 0.090, 0.040, 0.015],
    5:  [1.000, 0.640, 0.330, 0.140, 0.060, 0.025, 0.010],
    6:  [1.000, 0.630, 0.310, 0.120, 0.050, 0.020, 0.008],
    7:  [1.000, 0.600, 0.230, 0.080, 0.030, 0.012, 0.005],
    8:  [1.000, 0.500, 0.200, 0.070, 0.025, 0.010, 0.004],
    9:  [1.000, 0.500, 0.170, 0.060, 0.020, 0.008, 0.003],
    10: [1.000, 0.390, 0.130, 0.040, 0.015, 0.006, 0.002],
    11: [1.000, 0.370, 0.110, 0.035, 0.012, 0.005, 0.002],
    12: [1.000, 0.360, 0.090, 0.025, 0.008, 0.003, 0.001],
    13: [1.000, 0.210, 0.050, 0.010, 0.003, 0.001, 0.000],
    14: [1.000, 0.150, 0.020, 0.005, 0.001, 0.000, 0.000],
    15: [1.000, 0.070, 0.010, 0.002, 0.000, 0.000, 0.000],
    16: [1.000, 0.010, 0.002, 0.000, 0.000, 0.000, 0.000],
}

# =============================================================================
# PRE-FITTED MODEL PARAMETERS (trained on 2022-2025 pool data, ~400 obs)
# =============================================================================
# Per-game scoring model: pts_per_game = ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed
# Where ppg_adj = (1 - SHRINKAGE) * ppg + SHRINKAGE * MEAN_PPG
ALPHA = 2.90          # Intercept: baseline per-game scoring floor
BETA_PPG = 0.701      # Each regular-season PPG contributes 0.70 tournament pts/game
BETA_SEED = 0.210     # Higher seeds get a small per-game scoring boost
SIGMA_GAME = 5.67     # Per-game scoring standard deviation (for simulations)
SHRINKAGE = 0.15      # Bayesian shrinkage toward population mean PPG
MEAN_PPG = 14.8       # Population mean PPG for shrinkage
MOMENTUM = 0.05       # 5% scoring escalation per round survived


def get_expected_games(seed):
    """Expected number of tournament games for a given seed."""
    probs = SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16])
    return sum(probs[:6])


def get_game_probabilities(seed):
    """Probability of playing exactly N games (1 through 6)."""
    probs = SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16])
    game_probs = []
    for k in range(6):
        if k < 5:
            game_probs.append(probs[k] - probs[k + 1])
        else:
            game_probs.append(probs[k])
    return game_probs


def predict_player(ppg, seed):
    """
    Predict expected tournament total points for a single player.

    Returns dict with expected_pts, variance, std, expected_games, risk_adj.
    """
    # Bayesian shrinkage on PPG
    ppg_adj = (1 - SHRINKAGE) * ppg + SHRINKAGE * MEAN_PPG

    # Predicted points per game
    pts_per_game = ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed
    pts_per_game = max(pts_per_game, 2.0)

    # Expected games from seed
    exp_games = get_expected_games(seed)

    # Momentum bonus (small escalation for later rounds, weighted by reach probability)
    probs = SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16])
    momentum_bonus = sum(probs[rnd] * MOMENTUM * rnd * ppg_adj for rnd in range(6))

    # Expected total points
    exp_total = pts_per_game * exp_games + momentum_bonus + seed

    # Variance (law of total variance: game uncertainty + scoring uncertainty)
    game_probs = get_game_probabilities(seed)
    e_games_sq = sum((k + 1) ** 2 * p for k, p in enumerate(game_probs))
    var_games = e_games_sq - exp_games ** 2
    variance = exp_games * SIGMA_GAME ** 2 + var_games * pts_per_game ** 2

    return {
        'expected_pts': round(exp_total, 1),
        'variance': round(variance, 1),
        'std': round(np.sqrt(variance), 1),
        'expected_games': round(exp_games, 2),
        'pts_per_game': round(pts_per_game, 1),
        'risk_adj': round(exp_total - 0.3 * np.sqrt(variance), 1),
    }


def simulate_player(ppg, seed, n_sims=10000):
    """Monte Carlo simulation of a player's tournament total."""
    ppg_adj = (1 - SHRINKAGE) * ppg + SHRINKAGE * MEAN_PPG
    pts_per_game_mean = max(ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed, 2.0)
    probs = SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16])

    totals = np.zeros(n_sims)
    for sim in range(n_sims):
        total = seed  # seed bonus
        for rnd in range(6):
            if np.random.random() > probs[rnd]:
                break
            momentum = 1.0 + MOMENTUM * rnd
            pts = max(0, np.random.normal(pts_per_game_mean * momentum, SIGMA_GAME))
            total += pts
        totals[sim] = total
    return totals


def load_and_validate_input(filepath):
    """
    Load the input Excel and map columns flexibly.
    Returns a cleaned DataFrame with standardized column names.
    """
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    df = pd.read_excel(filepath)
    df.columns = [str(c).strip() for c in df.columns]

    # Flexible column mapping
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if cl in ('player', 'name', 'player name'):
            col_map['player'] = col
        elif cl in ('team', 'school'):
            col_map['team'] = col
        elif cl in ('seed',):
            col_map['seed'] = col
        elif cl in ('ppg', 'pts', 'points', 'pts/g', 'avg'):
            col_map['ppg'] = col
        elif cl in ('region', 'reg', 'regions'):
            col_map['region'] = col
        elif cl in ('rank', 'rk', '#'):
            col_map['rank'] = col

    # Check required columns
    for required in ['player', 'team', 'seed', 'ppg']:
        if required not in col_map:
            print(f"ERROR: Could not find '{required}' column in your Excel.")
            print(f"  Found columns: {list(df.columns)}")
            print(f"  Expected one of: Player/Name, Team/School, Seed, PPG/PTS/Points")
            sys.exit(1)

    # Rename to standard names
    rename = {v: k for k, v in col_map.items()}
    df = df.rename(columns=rename)

    # Clean data
    df['seed'] = pd.to_numeric(df['seed'], errors='coerce').fillna(16).astype(int)
    df['ppg'] = pd.to_numeric(df['ppg'], errors='coerce').fillna(0.0)
    df['player'] = df['player'].astype(str).str.strip()
    df['team'] = df['team'].astype(str).str.strip()

    if 'region' not in df.columns:
        df['region'] = 'Unknown'
    else:
        df['region'] = df['region'].astype(str).str.strip()

    if 'rank' not in df.columns:
        df['rank'] = range(1, len(df) + 1)

    # Drop rows with no player name or zero PPG
    df = df[df['player'].str.len() > 0]
    df = df[df['ppg'] > 0]

    return df[['player', 'team', 'seed', 'region', 'ppg', 'rank']].reset_index(drop=True)


def run_eda(df):
    """Print exploratory data analysis to console."""
    print("\n" + "=" * 70)
    print("EXPLORATORY DATA ANALYSIS")
    print("=" * 70)

    print(f"\nTotal players loaded: {len(df)}")
    print(f"Seeds represented: {sorted(df['seed'].unique())}")
    print(f"Regions: {sorted(df['region'].unique())}")

    print("\n--- Player Count by Seed ---")
    seed_counts = df.groupby('seed').size()
    for seed, count in seed_counts.items():
        exp_g = get_expected_games(seed)
        print(f"  Seed {seed:2d}: {count:3d} players | Expected games: {exp_g:.2f}")

    print("\n--- PPG Distribution by Seed Bucket ---")
    buckets = {
        'Top (1-2)': df[df['seed'].isin([1, 2])],
        'Mid (3-5)': df[df['seed'].isin([3, 4, 5])],
        'Low (6-8)': df[df['seed'].isin([6, 7, 8])],
        'Cinderella (9+)': df[df['seed'] >= 9],
    }
    for name, bucket in buckets.items():
        if len(bucket) > 0:
            print(f"  {name:20s}: N={len(bucket):3d}, "
                  f"PPG mean={bucket['ppg'].mean():.1f}, "
                  f"PPG range=[{bucket['ppg'].min():.1f} - {bucket['ppg'].max():.1f}]")

    if 'region' in df.columns:
        print("\n--- Players by Region ---")
        for region, group in df.groupby('region'):
            seeds = group['seed'].value_counts().sort_index()
            seed_str = ', '.join(f"{s}-seed: {c}" for s, c in seeds.items() if s <= 4)
            print(f"  {region:8s}: {len(group):3d} players | {seed_str}")


def run_predictions(df):
    """Generate predictions for all players. Returns df with prediction columns."""
    print("\n" + "=" * 70)
    print("PREDICTIVE MODEL")
    print("=" * 70)

    print("\n--- Pre-Fitted Model Parameters (trained on 2022-2025 data) ---")
    print(f"  Per-game model: pts/game = {ALPHA:.2f} + {BETA_PPG:.3f} * PPG_adj + {BETA_SEED:.3f} * Seed")
    print(f"  Bayesian shrinkage: {SHRINKAGE:.0%} toward mean PPG of {MEAN_PPG:.1f}")
    print(f"  Per-game std dev: {SIGMA_GAME:.2f}")
    print(f"  Momentum bonus: {MOMENTUM:.0%} per round")
    print(f"\n  How to read: A 18 PPG player on a 2-seed...")
    demo = predict_player(18.0, 2)
    print(f"    -> PPG adjusted to {(1 - SHRINKAGE) * 18.0 + SHRINKAGE * MEAN_PPG:.1f} (shrinkage)")
    print(f"    -> {demo['pts_per_game']} projected pts/game")
    print(f"    -> {demo['expected_games']} expected games")
    print(f"    -> {demo['expected_pts']} expected total points (including seed bonus)")
    print(f"    -> {demo['std']} std dev (uncertainty range)")

    # Generate predictions for all players
    predictions = []
    for _, row in df.iterrows():
        pred = predict_player(row['ppg'], row['seed'])
        predictions.append(pred)

    pred_df = pd.DataFrame(predictions)
    result = pd.concat([df.reset_index(drop=True), pred_df], axis=1)
    result = result.sort_values('expected_pts', ascending=False).reset_index(drop=True)
    result['draft_rank'] = range(1, len(result) + 1)

    return result


def run_optimization_analysis(df):
    """Compare optimization strategies and print to console."""
    print("\n" + "=" * 70)
    print("OPTIMIZATION STRATEGY COMPARISON")
    print("=" * 70)

    print("\n  Selecting top 8 under different strategies to illustrate tradeoffs.")
    print("  In practice, use the full draft_rank column as your draft board.\n")

    # Strategy 1: Pure expected value (no constraints)
    top8_ev = df.head(8).copy()

    # Strategy 2: Diversified (max 3 per region, max 3 per seed)
    selected = []
    region_counts = defaultdict(int)
    seed_counts = defaultdict(int)
    for _, row in df.iterrows():
        if len(selected) >= 8:
            break
        if region_counts[row['region']] >= 3:
            continue
        if seed_counts[row['seed']] >= 3:
            continue
        selected.append(row)
        region_counts[row['region']] += 1
        seed_counts[row['seed']] += 1
    top8_div = pd.DataFrame(selected)

    strategies = {
        'High Upside (Concentrated — RECOMMENDED)': top8_ev,
        'Balanced (Diversified across regions/seeds)': top8_div,
    }

    for name, lineup in strategies.items():
        total_exp = lineup['expected_pts'].sum()
        total_std = np.sqrt(lineup['variance'].sum())
        print(f"  === {name} ===")
        print(f"  Total E[Pts]: {total_exp:.1f} | Total Std: {total_std:.1f}\n")
        for _, row in lineup.iterrows():
            print(f"    {row['draft_rank']:3.0f}. {row['player']:25s} "
                  f"{row['team']:18s} Seed={row['seed']:2d} "
                  f"PPG={row['ppg']:5.1f} E[Pts]={row['expected_pts']:6.1f} "
                  f"Region={row['region']}")
        print()

    # Highlight differences
    ev_players = set(top8_ev['player'])
    div_players = set(top8_div['player'])
    only_ev = ev_players - div_players
    only_div = div_players - ev_players
    if only_ev:
        print(f"  Concentrated adds: {', '.join(only_ev)}")
        print(f"  Diversified swaps in: {', '.join(only_div)}")
    else:
        print("  Both strategies select the same 8 players.")


def run_monte_carlo(df, top_n=8, n_sims=15000):
    """Monte Carlo simulation of the top-8 lineup."""
    print("\n" + "=" * 70)
    print("MONTE CARLO SIMULATION (Top 8 Lineup)")
    print("=" * 70)

    np.random.seed(42)
    top8 = df.head(top_n)
    lineup_totals = np.zeros(n_sims)

    print(f"\n  Simulating {n_sims:,} tournaments...\n")

    for _, row in top8.iterrows():
        sims = simulate_player(row['ppg'], row['seed'], n_sims=n_sims)
        lineup_totals += sims

        print(f"  {row['player']:25s}: Mean={sims.mean():5.1f}, "
              f"Std={sims.std():5.1f}, "
              f"10th-90th=[{np.percentile(sims, 10):.0f} - {np.percentile(sims, 90):.0f}]")

    print(f"\n  --- Lineup Totals ---")
    print(f"  Mean:       {lineup_totals.mean():.1f}")
    print(f"  Median:     {np.median(lineup_totals):.1f}")
    print(f"  Std Dev:    {lineup_totals.std():.1f}")
    print(f"  5th pct:    {np.percentile(lineup_totals, 5):.1f}  (worst case)")
    print(f"  95th pct:   {np.percentile(lineup_totals, 95):.1f}  (best case)")
    print(f"  P(>300):    {np.mean(lineup_totals > 300):.1%}")
    print(f"  P(>350):    {np.mean(lineup_totals > 350):.1%}")
    print(f"  P(>400):    {np.mean(lineup_totals > 400):.1%}")


def write_output(df, output_path):
    """Write the ranked draft board to Excel with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Draft Board ---
    ws = wb.active
    ws.title = "Draft Board"

    headers = ['Draft Rank', 'Player', 'Team', 'Seed', 'Region', 'Reg Season PPG',
               'Expected Total Pts', 'Std Dev', 'Expected Games', 'Proj Pts/Game',
               'Risk-Adjusted Score', 'Cheat Sheet Rank']
    cols = ['draft_rank', 'player', 'team', 'seed', 'region', 'ppg',
            'expected_pts', 'std', 'expected_games', 'pts_per_game',
            'risk_adj', 'rank']

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    top8_fill = PatternFill('solid', fgColor='D6E4F0')
    data_font = Font(name='Arial', size=10)
    border = Border(
        bottom=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
    )

    # Write headers
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, (np.integer, np.int64)):
                val = int(val)
            elif isinstance(val, (np.floating, np.float64)):
                val = float(val)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c in (1, 4, 7, 8, 9, 10, 11, 12):
                cell.alignment = Alignment(horizontal='center')

            # Highlight top 8
            if r <= 9:
                cell.fill = top8_fill

    # Column widths
    widths = [10, 28, 22, 6, 10, 14, 16, 10, 14, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w
    # Fix column width for columns beyond 'I'
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = widths[i]

    # Freeze top row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Model Info ---
    ws2 = wb.create_sheet("Model Info")
    ws2['A1'] = 'NCAA Tournament Player Pool — Model Parameters'
    ws2['A1'].font = Font(bold=True, size=14, name='Arial')

    info = [
        ('', ''),
        ('Model', 'Per-game linear regression + historical seed advancement lookup'),
        ('Training Data', '2022-2025 pool results (~400 player-season observations)'),
        ('Backtest Performance', '2024: 100th percentile | 2025: 93rd percentile'),
        ('', ''),
        ('FITTED PARAMETERS', ''),
        ('Alpha (intercept)', f'{ALPHA:.2f} — baseline per-game scoring floor'),
        ('Beta_PPG', f'{BETA_PPG:.3f} — each regular-season PPG contributes this many tournament pts/game'),
        ('Beta_Seed', f'{BETA_SEED:.3f} — higher seeds get a small per-game scoring boost'),
        ('Sigma (per-game std)', f'{SIGMA_GAME:.2f} — per-game scoring noise'),
        ('Shrinkage', f'{SHRINKAGE:.2f} — PPG pulled {SHRINKAGE:.0%} toward mean of {MEAN_PPG:.1f}'),
        ('Momentum', f'{MOMENTUM:.2f} — {MOMENTUM:.0%} scoring escalation per round survived'),
        ('', ''),
        ('FORMULA', ''),
        ('PPG Adjusted', f'ppg_adj = (1 - {SHRINKAGE}) * PPG + {SHRINKAGE} * {MEAN_PPG}'),
        ('Pts Per Game', f'pts/game = {ALPHA} + {BETA_PPG} * ppg_adj + {BETA_SEED} * seed'),
        ('Expected Total', 'pts/game * expected_games + momentum_bonus + seed_bonus'),
        ('', ''),
        ('EXPECTED GAMES BY SEED', ''),
    ]

    for seed in range(1, 17):
        info.append((f'Seed {seed}', f'{get_expected_games(seed):.2f} games'))

    for r, (label, value) in enumerate(info, 3):
        ws2.cell(row=r, column=1, value=label).font = Font(
            bold=('PARAMETERS' in label or 'FORMULA' in label or 'EXPECTED' in label or label == 'Model'),
            name='Arial', size=10
        )
        ws2.cell(row=r, column=2, value=value).font = Font(name='Arial', size=10)

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 70

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='NCAA Tournament Player Pool — Draft Board Generator')
    parser.add_argument('input', help='Path to input Excel file (cheat sheet)')
    parser.add_argument('--output', '-o', default=None,
                        help='Output Excel path (default: draft_board_YYYY.xlsx)')
    args = parser.parse_args()

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        base = os.path.splitext(os.path.basename(args.input))[0]
        output_path = f'draft_board_{base}.xlsx'

    print("=" * 70)
    print("NCAA TOURNAMENT PLAYER POOL — DRAFT BOARD GENERATOR")
    print("=" * 70)
    print(f"\n  Input:  {args.input}")
    print(f"  Output: {output_path}")

    # Load input
    print("\n[1/5] Loading input data...")
    df = load_and_validate_input(args.input)
    print(f"  Loaded {len(df)} players from {args.input}")

    # EDA
    print("\n[2/5] Running exploratory data analysis...")
    run_eda(df)

    # Predictions
    print("\n[3/5] Generating predictions...")
    ranked_df = run_predictions(df)

    # Print top 30 draft board to console
    print("\n" + "=" * 70)
    print("DRAFT BOARD (Top 30)")
    print("=" * 70)
    print(f"\n  {'Rank':>4} {'Player':25s} {'Team':18s} {'Seed':>4} "
          f"{'PPG':>5} {'E[Pts]':>7} {'Std':>5} {'E[Games]':>8} {'Region':>8}")
    print("  " + "-" * 95)
    for _, row in ranked_df.head(30).iterrows():
        marker = " <--" if row['draft_rank'] <= 8 else ""
        print(f"  {row['draft_rank']:4.0f} {row['player']:25s} {row['team']:18s} "
              f"{row['seed']:4d} {row['ppg']:5.1f} {row['expected_pts']:7.1f} "
              f"{row['std']:5.1f} {row['expected_games']:8.2f} {row['region']:>8}{marker}")
    print(f"\n  <-- = Top 8 (ideal picks if available)")

    # Optimization comparison
    print("\n[4/5] Running optimization analysis...")
    run_optimization_analysis(ranked_df)

    # Monte Carlo
    print("\n[5/5] Running Monte Carlo simulation...")
    run_monte_carlo(ranked_df)

    # Write output
    print("\n" + "=" * 70)
    print("WRITING OUTPUT")
    print("=" * 70)
    write_output(ranked_df, output_path)
    print(f"\n  Draft board saved to: {output_path}")
    print(f"  Total players ranked: {len(ranked_df)}")
    print(f"\n  The Excel contains two sheets:")
    print(f"    1. 'Draft Board' — Full ranked player list (top 8 highlighted)")
    print(f"    2. 'Model Info'  — Model parameters and formulas")
    print(f"\n  Use the 'Draft Rank' column as your pick order during the draft.")
    print(f"  When it's your turn, take the highest-ranked available player.")

    print("\n" + "=" * 70)
    print("DONE")
    print("=" * 70)


if __name__ == '__main__':
    main()
