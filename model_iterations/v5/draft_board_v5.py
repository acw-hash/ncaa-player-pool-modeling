#!/usr/bin/env python3
"""
NCAA Tournament Player Pool — Draft Board Generator (V5)
==========================================================
Methodological improvements over V4:
  5. BARTHAG advancement adjustment: Barttorvik's composite win probability
     replaces AdjEM as the primary team-quality advancement predictor.
     Validated via exhaustive 10-method analysis (logistic, RF, elastic net,
     bootstrap, etc.) on pre-tournament data.
  6. Three-point rate (3PR) advancement adjustment: The ONLY feature with a
     statistically reliable bootstrap confidence interval [0.11, 0.43] for
     predicting tournament advancement beyond seed. Teams that shoot more
     threes advance further — confirmed across all 10 analytical methods.

Carried forward from V4:
  1. Team-correlated advancement (players on same team advance together)
  2. Per-seed-bucket shrinkage (empirical Bayes calibrated)
  3. Win-probability evaluation (full pool competition simulation)
  4. Bootstrap-augmented CV with 1-SE rule (minimal PPG+Seed model)

USAGE:
    python draft_board_v5.py input.xlsx
    python draft_board_v5.py input.xlsx --output my_draft_board.xlsx
    python draft_board_v5.py input.xlsx --risk 0.3

INPUT:
    An Excel file with columns (case-insensitive, flexible naming):

    REQUIRED:
        - Player (or Name)
        - Team (or School)
        - Seed
        - PPG (or PTS, Points)
        - Region (optional)
        - Rank (optional)

    OPTIONAL (team-level + player usage — used for advancement & tiebreakers):
        - BARTHAG (Barttorvik win probability, e.g., 0.9824)
        - ThreeRate (or 3PR, Three-point attempt rate, e.g., 38.5)
        - AdjEM (KenPom Adjusted Efficiency Margin — fallback if no BARTHAG)
        - AdjO  (KenPom Adjusted Offensive Efficiency)
        - Pace  (Possessions per 40 min)
        - Q1_Wins (Quadrant 1 wins)
        - SOR_Rank (Strength of Record ranking)
        - Usage (or USG, Usage_Pct — player usage rate)

    If optional columns are missing, the model falls back gracefully.
    Fully backward-compatible with V1/V2/V3 input format.

OUTPUT:
    An Excel file with ranked draft board + model diagnostics + team metrics.

MODEL (V5 — BARTHAG + 3PR Advancement):
    Per-game scoring: Ridge(PPG + Seed), alpha=50 (1-SE rule validated)
    Enhanced tiebreaker: Ridge(PPG + Seed + Pace + Usage + AdjO), alpha=50
    Advancement: Historical seed probs + BARTHAG (gamma=0.08) + 3PR (gamma=0.05)
    Shrinkage: Per-seed-bucket (0.15/0.60/0.20/0.60) — empirical Bayes calibrated
    Monte Carlo: Team-correlated advancement + Start-5 rule + heteroscedastic variance
    See NCAA_Pool_Enhanced_Analysis_Report_V5.md for full methodology.
"""

import pandas as pd
import numpy as np
import argparse
import sys
import os
from collections import defaultdict

# =============================================================================
# HISTORICAL SEED ADVANCEMENT PROBABILITIES (1985-2024, ~40 years of data)
# Each list: [P(play R64), P(reach R32), P(reach S16), P(reach E8),
#             P(reach F4), P(reach NCG), P(win NCG)]
# =============================================================================
SEED_ADVANCE_PROBS = {
    1:  [1.000, 0.993, 0.850, 0.600, 0.390, 0.220, 0.130],
    2:  [1.000, 0.940, 0.670, 0.400, 0.220, 0.110, 0.060],
    3:  [1.000, 0.850, 0.510, 0.250, 0.120, 0.050, 0.020],
    4:  [1.000, 0.790, 0.430, 0.200, 0.090, 0.040, 0.015],
    5:  [1.000, 0.640, 0.330, 0.140, 0.060, 0.025, 0.010],
    6:  [1.000, 0.630, 0.310, 0.120, 0.050, 0.020, 0.008],
    7:  [1.000, 0.600, 0.230, 0.080, 0.030, 0.012, 0.005],
    8:  [1.000, 0.500, 0.200, 0.070, 0.025, 0.010, 0.004],
    9:  [1.000, 0.500, 0.170, 0.060, 0.020, 0.008, 0.003],
    10: [1.000, 0.390, 0.130, 0.040, 0.015, 0.006, 0.002],
    11: [1.000, 0.370, 0.110, 0.035, 0.012, 0.005, 0.002],
    12: [1.000, 0.360, 0.090, 0.025, 0.008, 0.003, 0.001],
    13: [1.000, 0.210, 0.050, 0.010, 0.003, 0.001, 0.000],
    14: [1.000, 0.150, 0.020, 0.005, 0.001, 0.000, 0.000],
    15: [1.000, 0.070, 0.010, 0.002, 0.000, 0.000, 0.000],
    16: [1.000, 0.010, 0.002, 0.000, 0.000, 0.000, 0.000],
}

# =============================================================================
# PRE-FITTED MODEL PARAMETERS (V4 — Bootstrap CV + 1-SE Rule)
# Trained on 2022-2025, 177 observations, Ridge alpha=50
# All hyperparameters selected via bootstrap-augmented LOYO-CV with 1-SE rule
# =============================================================================

# --- Primary Scoring Model (Ridge=50, minimal features, 1-SE validated) ---
# pts/game = ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed
ALPHA = -1.813          # Intercept
BETA_PPG = 1.144        # Regular-season PPG effect
BETA_SEED = -0.844      # Seed effect on per-game scoring

# --- Enhanced Tiebreaker Model (Ridge=50, full features) ---
# Used when team metrics are available; within 1-SE of minimal
ENHANCED_ALPHA = 2.328
ENHANCED_BETA_PPG = 0.864
ENHANCED_BETA_SEED = -0.867
ENHANCED_BETA_PACE = 0.199
ENHANCED_BETA_USAGE = 0.961
ENHANCED_BETA_ADJO = 0.009

# --- Per-Seed-Bucket Shrinkage (V4 — Empirical Bayes calibrated) ---
# Higher seeds have more reliable PPG (stronger schedule, more games)
# Lower seeds have inflated PPG from weaker competition -> more shrinkage
SHRINKAGE_TOP = 0.15        # Seeds 1-2: minimal shrinkage (reliable PPG)
SHRINKAGE_MID = 0.60        # Seeds 3-5: heavy shrinkage (PPG less predictive)
SHRINKAGE_LOW = 0.20        # Seeds 6-8: moderate shrinkage
SHRINKAGE_CINDERELLA = 0.60 # Seeds 9+: heavy shrinkage (weak schedule inflates PPG)
MEAN_PPG = 14.8             # Population mean PPG for shrinkage target

# --- Heteroscedastic variance by seed bucket ---
SIGMA_SEEDS_1_2 = 3.81
SIGMA_SEEDS_3_5 = 3.82
SIGMA_SEEDS_6_8 = 2.99
SIGMA_SEEDS_9P = 6.52
SIGMA_GLOBAL = 4.06

# --- Normalization constants (from training data) ---
PACE_MEAN = 68.67
PACE_STD = 2.60
USAGE_MEAN = 22.79
USAGE_STD = 4.23
ADJO_MEAN = 115.84
ADJO_STD = 4.45

# --- Advancement Adjustment (V5: BARTHAG + Three-Point Rate) ---
# BARTHAG: Barttorvik composite win probability — best single team quality metric.
#   Validated via 10-method exhaustive analysis. Strongest threshold effect of any
#   feature: top-quartile teams play 3.9 games vs 2.3 (Δ=1.6 games).
# Three-Point Rate: ONLY feature with reliable bootstrap CI [0.11, 0.43].
#   Teams shooting more threes advance further — higher-variance offense + modern style.
# AdjEM: Replaced by BARTHAG (r=0.95 correlation; BARTHAG is strictly superior composite).
GAMMA_BARTHAG = 0.08   # BARTHAG effect on log-odds of advancing per round
GAMMA_3PR = 0.05       # Three-point rate effect on advancement
BARTHAG_MEAN = 0.9226  # Training data mean
BARTHAG_STD = 0.0717   # Training data std
THREE_RATE_MEAN = 38.41
THREE_RATE_STD = 5.20
# Legacy AdjEM constants (used only as fallback when BARTHAG unavailable)
GAMMA_EM = 0.08
ADJEM_MEAN = 21.00
ADJEM_STD = 6.02

# --- Fallback: Base Model (same as primary — minimal is the 1-SE winner) ---
V1_ALPHA = -1.813
V1_BETA_PPG = 1.144
V1_BETA_SEED = -0.844
V1_SIGMA = 4.06


def get_shrinkage_for_seed(seed):
    """Get the per-seed-bucket shrinkage value (V4 improvement)."""
    if seed <= 2:
        return SHRINKAGE_TOP
    elif seed <= 5:
        return SHRINKAGE_MID
    elif seed <= 8:
        return SHRINKAGE_LOW
    else:
        return SHRINKAGE_CINDERELLA


def get_sigma_for_seed(seed):
    """Get the heteroscedastic per-game scoring std for this seed."""
    if seed <= 2:
        return SIGMA_SEEDS_1_2
    elif seed <= 5:
        return SIGMA_SEEDS_3_5
    elif seed <= 8:
        return SIGMA_SEEDS_6_8
    else:
        return SIGMA_SEEDS_9P


def get_base_expected_games(seed):
    """Expected number of tournament games using base seed probabilities."""
    probs = SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16])
    return sum(probs[:6])


def get_adjusted_advance_probs(seed, barthag=None, three_rate=None, adj_em=None):
    """
    Get advancement probabilities, adjusted by BARTHAG + three-point rate (V5).
    Falls back to AdjEM if BARTHAG unavailable (backward compatible).
    """
    probs = list(SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16]))

    # Compute total logit adjustment from all available features
    total_adj = 0.0

    if barthag is not None and not np.isnan(barthag):
        total_adj += GAMMA_BARTHAG * (barthag - BARTHAG_MEAN) / BARTHAG_STD
    elif adj_em is not None and not np.isnan(adj_em):
        # Fallback to AdjEM if no BARTHAG
        total_adj += GAMMA_EM * (adj_em - ADJEM_MEAN) / ADJEM_STD

    if three_rate is not None and not np.isnan(three_rate):
        total_adj += GAMMA_3PR * (three_rate - THREE_RATE_MEAN) / THREE_RATE_STD

    if abs(total_adj) > 0.001:
        adjusted = [probs[0]]
        for r in range(1, len(probs)):
            base_p = max(min(probs[r], 0.999), 0.001)
            logit = np.log(base_p / (1 - base_p))
            logit_adj = logit + total_adj
            adj_p = 1 / (1 + np.exp(-logit_adj))
            adj_p = min(adj_p, adjusted[-1])
            adjusted.append(adj_p)
        return adjusted
    return probs


def get_conditional_probs(probs):
    """Convert cumulative advancement probs to conditional (round-by-round)."""
    cond = [1.0]
    for rnd in range(1, 6):
        if probs[rnd - 1] > 0:
            cond.append(probs[rnd] / probs[rnd - 1])
        else:
            cond.append(0.0)
    return cond


def get_game_probabilities(probs):
    """Probability of playing exactly N games (1 through 6) from advancement probs."""
    game_probs = []
    for k in range(6):
        if k < 5:
            game_probs.append(probs[k] - probs[k + 1])
        else:
            game_probs.append(probs[k])
    return game_probs


def predict_player_enhanced(ppg, seed, adj_em=None, adj_o=None, pace=None,
                            usage=None, barthag=None, three_rate=None,
                            use_enhanced=True):
    """
    Predict expected tournament total points using the V5 model.

    V5 changes:
    - BARTHAG + three-point rate in advancement model (replaces AdjEM-only)
    Carried from V4:
    - Per-seed-bucket shrinkage (empirical Bayes calibrated)
    - 1-SE validated minimal model as primary
    - Enhanced model as tiebreaker when team metrics available
    """
    # Seed-dependent Bayesian shrinkage
    lam = get_shrinkage_for_seed(seed)
    ppg_adj = (1 - lam) * ppg + lam * MEAN_PPG

    # Determine if we have team metrics for enhanced model
    has_team_metrics = (use_enhanced and pace is not None and not np.isnan(pace)
                        and usage is not None and not np.isnan(usage))

    if has_team_metrics:
        # Enhanced tiebreaker model (within 1-SE, provides marginal info)
        pace_norm = (pace - PACE_MEAN) / PACE_STD
        usage_norm = (usage - USAGE_MEAN) / USAGE_STD
        adjo_norm = ((adj_o if adj_o is not None and not np.isnan(adj_o) else ADJO_MEAN)
                     - ADJO_MEAN) / ADJO_STD

        pts_per_game = (ENHANCED_ALPHA + ENHANCED_BETA_PPG * ppg_adj +
                       ENHANCED_BETA_SEED * seed + ENHANCED_BETA_PACE * pace_norm +
                       ENHANCED_BETA_USAGE * usage_norm + ENHANCED_BETA_ADJO * adjo_norm)
        sigma = get_sigma_for_seed(seed)
        model_used = 'enhanced'
    else:
        # Primary model (1-SE validated)
        pts_per_game = ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed
        sigma = get_sigma_for_seed(seed)
        model_used = 'base'

    pts_per_game = max(pts_per_game, 2.0)

    # V5: Adjusted advancement using BARTHAG + three-point rate
    adv_probs = get_adjusted_advance_probs(seed, barthag=barthag,
                                            three_rate=three_rate, adj_em=adj_em)
    exp_games = sum(adv_probs[:6])

    # Expected total points (no momentum — V3/V4)
    exp_total = pts_per_game * exp_games + seed

    # Variance (law of total variance)
    game_probs = get_game_probabilities(adv_probs)
    e_games_sq = sum((k + 1) ** 2 * p for k, p in enumerate(game_probs))
    var_games = e_games_sq - exp_games ** 2
    variance = exp_games * sigma ** 2 + var_games * pts_per_game ** 2

    return {
        'expected_pts': round(exp_total, 1),
        'variance': round(variance, 1),
        'std': round(np.sqrt(max(variance, 0)), 1),
        'expected_games': round(exp_games, 2),
        'pts_per_game': round(pts_per_game, 1),
        'risk_adj': round(exp_total - 0.3 * np.sqrt(max(variance, 0)), 1),
        'model_used': model_used,
        'shrinkage': round(lam, 2),
    }


def simulate_player_enhanced(ppg, seed, adj_em=None, adj_o=None, pace=None,
                              usage=None, barthag=None, three_rate=None,
                              n_sims=10000):
    """
    Monte Carlo simulation of a single player's tournament total.
    Uses corrected conditional probabilities and heteroscedastic variance.
    V5: advancement adjusted by BARTHAG + three-point rate.
    """
    lam = get_shrinkage_for_seed(seed)
    ppg_adj = (1 - lam) * ppg + lam * MEAN_PPG

    has_team = (pace is not None and not np.isnan(pace)
                and usage is not None and not np.isnan(usage))

    if has_team:
        pace_norm = (pace - PACE_MEAN) / PACE_STD
        usage_norm = (usage - USAGE_MEAN) / USAGE_STD
        adjo_norm = ((adj_o if adj_o is not None and not np.isnan(adj_o) else ADJO_MEAN)
                     - ADJO_MEAN) / ADJO_STD
        pts_per_game_mean = max(ENHANCED_ALPHA + ENHANCED_BETA_PPG * ppg_adj +
                               ENHANCED_BETA_SEED * seed +
                               ENHANCED_BETA_PACE * pace_norm +
                               ENHANCED_BETA_USAGE * usage_norm +
                               ENHANCED_BETA_ADJO * adjo_norm, 2.0)
    else:
        pts_per_game_mean = max(ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed, 2.0)

    sigma = get_sigma_for_seed(seed)
    probs = get_adjusted_advance_probs(seed, barthag=barthag,
                                        three_rate=three_rate, adj_em=adj_em)
    cond_probs = get_conditional_probs(probs)

    totals = np.zeros(n_sims)
    for sim in range(n_sims):
        total = seed
        for rnd in range(6):
            if rnd > 0 and np.random.random() > cond_probs[rnd]:
                break
            pts = max(0, np.random.normal(pts_per_game_mean, sigma))
            total += pts
        totals[sim] = total
    return totals


def simulate_lineup_team_correlated(players_data, n_sims=15000):
    """
    Monte Carlo simulation of an 8-player lineup WITH:
    - V4: Team-correlated advancement (players on same team advance together)
    - Start-5 rule (only 5 highest-scoring active players count per round)
    - Heteroscedastic variance by seed bucket

    Parameters:
    -----------
    players_data : list of dicts with keys:
        team, seed, ppg, adj_em, adj_o, pace, usage (optional team metrics)

    Returns array of lineup totals across simulations.
    """
    # Group players by team for correlated advancement
    teams = {}
    for i, p in enumerate(players_data):
        team = p['team']
        if team not in teams:
            teams[team] = {
                'seed': p['seed'],
                'adj_em': p.get('adj_em'),
                'barthag': p.get('barthag'),
                'three_rate': p.get('three_rate'),
                'player_indices': [],
            }
        teams[team]['player_indices'].append(i)

    # Precompute per-team conditional advancement probabilities (V5: BARTHAG + 3PR)
    team_cond_probs = {}
    for team, info in teams.items():
        probs = get_adjusted_advance_probs(info['seed'], barthag=info['barthag'],
                                            three_rate=info['three_rate'],
                                            adj_em=info['adj_em'])
        team_cond_probs[team] = get_conditional_probs(probs)

    # Precompute per-player scoring parameters
    player_params = []
    for p in players_data:
        lam = get_shrinkage_for_seed(p['seed'])
        ppg_adj = (1 - lam) * p['ppg'] + lam * MEAN_PPG

        pace_val = p.get('pace')
        usage_val = p.get('usage')
        has_team = (pace_val is not None and not np.isnan(pace_val)
                    and usage_val is not None and not np.isnan(usage_val))

        if has_team:
            pace_norm = (pace_val - PACE_MEAN) / PACE_STD
            usage_norm = (usage_val - USAGE_MEAN) / USAGE_STD
            adjo = p.get('adj_o')
            adjo = adjo if adjo is not None and not np.isnan(adjo) else ADJO_MEAN
            adjo_norm = (adjo - ADJO_MEAN) / ADJO_STD
            pts_mean = max(ENHANCED_ALPHA + ENHANCED_BETA_PPG * ppg_adj +
                          ENHANCED_BETA_SEED * p['seed'] +
                          ENHANCED_BETA_PACE * pace_norm +
                          ENHANCED_BETA_USAGE * usage_norm +
                          ENHANCED_BETA_ADJO * adjo_norm, 2.0)
        else:
            pts_mean = max(ALPHA + BETA_PPG * ppg_adj + BETA_SEED * p['seed'], 2.0)

        player_params.append({
            'seed': p['seed'],
            'team': p['team'],
            'pts_mean': pts_mean,
            'sigma': get_sigma_for_seed(p['seed']),
        })

    lineup_totals = np.zeros(n_sims)
    seed_total = sum(pp['seed'] for pp in player_params)
    n_players = len(player_params)

    for sim in range(n_sims):
        # STEP 1: Simulate team advancement ONCE per team (V4 key improvement)
        team_last_round = {}
        for team, info in teams.items():
            cond = team_cond_probs[team]
            last_rnd = 0
            for rnd in range(1, 6):
                if np.random.random() <= cond[rnd]:
                    last_rnd = rnd
                else:
                    break
            team_last_round[team] = last_rnd

        # STEP 2: Score each round with Start-5 rule
        total = seed_total
        for rnd in range(6):
            active_scores = []
            for i in range(n_players):
                if team_last_round[player_params[i]['team']] >= rnd:
                    score = max(0, np.random.normal(
                        player_params[i]['pts_mean'], player_params[i]['sigma']))
                    active_scores.append(score)

            if not active_scores:
                break
            active_scores.sort(reverse=True)
            total += sum(active_scores[:5])  # Start-5 rule

        lineup_totals[sim] = total

    return lineup_totals


def draft_opponent_lineup(all_players, strategy, rng, n_picks=8):
    """Generate a plausible opponent lineup for win-probability simulation."""
    available = list(all_players)

    if strategy == 'cheatsheet':
        # Pick by original cheat sheet rank (most common strategy)
        start = rng.randint(0, min(5, len(available)))
        lineup = []
        for p in available[start:]:
            if len(lineup) >= n_picks:
                break
            if rng.random() > 0.15:  # 85% chance they take each player
                lineup.append(p)
        while len(lineup) < n_picks and available:
            lineup.append(available[rng.randint(0, len(available))])
        return lineup[:n_picks]

    elif strategy == 'ppg_greedy':
        available_sorted = sorted(available, key=lambda p: -p['ppg'])
        start = rng.randint(0, min(3, len(available_sorted)))
        return available_sorted[start:start + n_picks]

    elif strategy == 'random_smart':
        top_pool = available[:min(60, len(available))]
        rng.shuffle(top_pool)
        return top_pool[:n_picks]

    else:
        rng.shuffle(available)
        return available[:n_picks]


def simulate_win_probability(model_lineup, all_player_data, n_competitors=18,
                              n_sims=3000):
    """
    V4: Simulate full pool competitions to estimate P(model lineup wins).
    Uses team-correlated advancement for both model and opponent lineups.

    Returns P(win), and the distribution of model scores for diagnostics.
    """
    rng = np.random.RandomState(42)
    wins = 0
    model_scores_all = []
    strategies = ['cheatsheet', 'ppg_greedy', 'random_smart']

    for sim in range(n_sims):
        # Score model lineup (team-correlated)
        model_score = simulate_lineup_team_correlated(model_lineup, n_sims=1)[0]
        model_scores_all.append(model_score)

        # Score opponent lineups
        best_opp = 0
        for c in range(n_competitors - 1):
            strat = strategies[c % len(strategies)]
            opp_players = draft_opponent_lineup(all_player_data, strat, rng)
            opp_score = simulate_lineup_team_correlated(opp_players, n_sims=1)[0]
            best_opp = max(best_opp, opp_score)

        if model_score > best_opp:
            wins += 1

    return wins / n_sims, np.array(model_scores_all)


def load_and_validate_input(filepath):
    """Load the input Excel and map columns flexibly."""
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    df = pd.read_excel(filepath)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {}
    for col in df.columns:
        cl = col.lower().replace('_', '').replace(' ', '')
        if cl in ('player', 'name', 'playername'):
            col_map['player'] = col
        elif cl in ('team', 'school'):
            col_map['team'] = col
        elif cl in ('seed',):
            col_map['seed'] = col
        elif cl in ('ppg', 'pts', 'points', 'pts/g', 'avg'):
            col_map['ppg'] = col
        elif cl in ('region', 'reg', 'regions'):
            col_map['region'] = col
        elif cl in ('rank', 'rk', '#'):
            col_map['rank'] = col
        elif cl in ('adjem', 'adjeffmargin', 'kenpomem', 'em'):
            col_map['adj_em'] = col
        elif cl in ('adjo', 'adjoffeff', 'adjoe', 'offeff', 'offrating'):
            col_map['adj_o'] = col
        elif cl in ('pace', 'possessions', 'tempo'):
            col_map['pace'] = col
        elif cl in ('q1wins', 'q1', 'quad1wins', 'quad1'):
            col_map['q1_wins'] = col
        elif cl in ('sorrank', 'sor', 'netrank', 'net', 'strengthofrecord'):
            col_map['sor_rank'] = col
        elif cl in ('usage', 'usg', 'usagepct', 'usg%', 'usagerate'):
            col_map['usage'] = col
        elif cl in ('barthag', 'barthagrk', 'winprob', 'btwp'):
            col_map['barthag'] = col
        elif cl in ('threerate', '3pr', 'threeptrate', '3ptrate', 'three_rate', 'threerate'):
            col_map['three_rate'] = col

    for required in ['player', 'team', 'seed', 'ppg']:
        if required not in col_map:
            print(f"ERROR: Could not find '{required}' column in your Excel.")
            print(f"  Found columns: {list(df.columns)}")
            sys.exit(1)

    rename = {v: k for k, v in col_map.items()}
    df = df.rename(columns=rename)

    df['seed'] = pd.to_numeric(df['seed'], errors='coerce').fillna(16).astype(int)
    df['ppg'] = pd.to_numeric(df['ppg'], errors='coerce').fillna(0.0)
    df['player'] = df['player'].astype(str).str.strip()
    df['team'] = df['team'].astype(str).str.strip()

    if 'region' not in df.columns:
        df['region'] = 'Unknown'
    else:
        df['region'] = df['region'].astype(str).str.strip()

    if 'rank' not in df.columns:
        df['rank'] = range(1, len(df) + 1)

    for opt_col in ['adj_em', 'adj_o', 'pace', 'q1_wins', 'sor_rank', 'usage',
                     'barthag', 'three_rate']:
        if opt_col in df.columns:
            df[opt_col] = pd.to_numeric(df[opt_col], errors='coerce')
        else:
            df[opt_col] = np.nan

    df = df[df['player'].str.len() > 0]
    df = df[df['ppg'] > 0]

    v5_cols_found = [c for c in ['barthag', 'three_rate', 'adj_em', 'adj_o', 'pace',
                                  'q1_wins', 'sor_rank', 'usage']
                     if df[c].notna().any()]
    if v5_cols_found:
        print(f"  Enhanced columns detected: {v5_cols_found}")
        if 'barthag' in v5_cols_found:
            print(f"  -> V5 advancement model: BARTHAG + 3PR (primary)")
        elif 'adj_em' in v5_cols_found:
            print(f"  -> V5 advancement model: AdjEM fallback (no BARTHAG)")
        else:
            print(f"  -> Advancement: seed-only (no team quality metrics)")
        if any(c in v5_cols_found for c in ['pace', 'usage']):
            print(f"  -> Scoring model: enhanced tiebreaker active")
    else:
        print(f"  No enhanced columns detected. Using primary model (PPG + Seed).")

    return df[['player', 'team', 'seed', 'region', 'ppg', 'rank',
               'adj_em', 'adj_o', 'pace', 'q1_wins', 'sor_rank', 'usage',
               'barthag', 'three_rate']].reset_index(drop=True)


def run_eda(df):
    """Print exploratory data analysis to console."""
    print("\n" + "=" * 70)
    print("EXPLORATORY DATA ANALYSIS")
    print("=" * 70)

    print(f"\nTotal players loaded: {len(df)}")
    print(f"Seeds represented: {sorted(df['seed'].unique())}")
    print(f"Regions: {sorted(df['region'].unique())}")

    print("\n--- Player Count by Seed ---")
    seed_counts = df.groupby('seed').size()
    for seed, count in seed_counts.items():
        exp_g = get_base_expected_games(seed)
        lam = get_shrinkage_for_seed(seed)
        print(f"  Seed {seed:2d}: {count:3d} players | E[games]: {exp_g:.2f} | "
              f"Shrinkage: {lam:.0%}")

    print("\n--- PPG Distribution by Seed Bucket ---")
    buckets = {
        'Top (1-2)': df[df['seed'].isin([1, 2])],
        'Mid (3-5)': df[df['seed'].isin([3, 4, 5])],
        'Low (6-8)': df[df['seed'].isin([6, 7, 8])],
        'Cinderella (9+)': df[df['seed'] >= 9],
    }
    for name, bucket in buckets.items():
        if len(bucket) > 0:
            print(f"  {name:20s}: N={len(bucket):3d}, "
                  f"PPG mean={bucket['ppg'].mean():.1f}, "
                  f"PPG range=[{bucket['ppg'].min():.1f} - {bucket['ppg'].max():.1f}]")

    if 'region' in df.columns:
        print("\n--- Players by Region ---")
        for region, group in df.groupby('region'):
            seeds = group['seed'].value_counts().sort_index()
            seed_str = ', '.join(f"{s}-seed: {c}" for s, c in seeds.items() if s <= 4)
            print(f"  {region:8s}: {len(group):3d} players | {seed_str}")


def run_predictions(df):
    """Generate predictions for all players."""
    print("\n" + "=" * 70)
    print("PREDICTIVE MODEL (V5)")
    print("=" * 70)

    print("\n--- V5 Model Parameters ---")
    print(f"  Primary model (1-SE validated): pts/game = {ALPHA:.3f} + "
          f"{BETA_PPG:.3f}*PPG_adj + {BETA_SEED:.3f}*Seed")
    print(f"  Shrinkage: Top(1-2)={SHRINKAGE_TOP:.0%}, Mid(3-5)={SHRINKAGE_MID:.0%}, "
          f"Low(6-8)={SHRINKAGE_LOW:.0%}, Cinderella(9+)={SHRINKAGE_CINDERELLA:.0%}")
    print(f"  Per-game std: Seeds 1-2={SIGMA_SEEDS_1_2}, 3-5={SIGMA_SEEDS_3_5}, "
          f"6-8={SIGMA_SEEDS_6_8}, 9+={SIGMA_SEEDS_9P}")

    has_enhanced = df['adj_em'].notna().any() or df['barthag'].notna().any()
    has_barthag = df['barthag'].notna().any()
    has_3pr = df['three_rate'].notna().any()
    if has_enhanced:
        print(f"\n  Enhanced tiebreaker model active (team metrics detected)")
        if has_barthag:
            print(f"  V5 advancement: BARTHAG (γ={GAMMA_BARTHAG})" +
                  (f" + 3PR (γ={GAMMA_3PR})" if has_3pr else ""))
        else:
            print(f"  V5 advancement: AdjEM fallback (γ={GAMMA_EM})" +
                  (f" + 3PR (γ={GAMMA_3PR})" if has_3pr else ""))

    demo_ppg, demo_seed = 18.0, 2
    demo = predict_player_enhanced(demo_ppg, demo_seed)
    lam_demo = get_shrinkage_for_seed(demo_seed)
    ppg_adj_demo = (1 - lam_demo) * demo_ppg + lam_demo * MEAN_PPG
    print(f"\n  Example: {demo_ppg} PPG player on {demo_seed}-seed...")
    print(f"    -> PPG shrunk to {ppg_adj_demo:.1f} (λ={lam_demo:.0%} for seed {demo_seed})")
    print(f"    -> {demo['pts_per_game']} projected pts/game")
    print(f"    -> {demo['expected_games']} expected games")
    print(f"    -> {demo['expected_pts']} expected total points")
    print(f"    -> {demo['std']} std dev")

    # Generate predictions
    predictions = []
    for _, row in df.iterrows():
        pred = predict_player_enhanced(
            row['ppg'], row['seed'],
            adj_em=row['adj_em'], adj_o=row['adj_o'],
            pace=row['pace'], usage=row['usage'],
            barthag=row['barthag'], three_rate=row['three_rate'])
        predictions.append(pred)

    pred_df = pd.DataFrame(predictions)
    result = pd.concat([df.reset_index(drop=True), pred_df], axis=1)
    result = result.sort_values('expected_pts', ascending=False).reset_index(drop=True)
    result['draft_rank'] = range(1, len(result) + 1)

    return result


def run_optimization_analysis(df):
    """Compare optimization strategies."""
    print("\n" + "=" * 70)
    print("OPTIMIZATION STRATEGY COMPARISON")
    print("=" * 70)

    print("\n  Selecting top 8 under different strategies.")
    print("  V4: Monte Carlo uses team-correlated advancement.\n")

    # Strategy 1: Pure expected value
    top8_ev = df.head(8).copy()

    # Strategy 2: Diversified
    selected = []
    region_counts = defaultdict(int)
    seed_counts = defaultdict(int)
    for _, row in df.iterrows():
        if len(selected) >= 8:
            break
        if region_counts[row['region']] >= 3:
            continue
        if seed_counts[row['seed']] >= 3:
            continue
        selected.append(row)
        region_counts[row['region']] += 1
        seed_counts[row['seed']] += 1
    top8_div = pd.DataFrame(selected)

    strategies = {
        'High Upside (Concentrated — RECOMMENDED)': top8_ev,
        'Balanced (Diversified across regions/seeds)': top8_div,
    }

    for name, lineup in strategies.items():
        total_exp = lineup['expected_pts'].sum()
        total_std = np.sqrt(lineup['variance'].sum())
        print(f"  === {name} ===")
        print(f"  Total E[Pts]: {total_exp:.1f} | Total Std: {total_std:.1f}\n")
        for _, row in lineup.iterrows():
            print(f"    {row['draft_rank']:3.0f}. {row['player']:25s} "
                  f"{row['team']:18s} Seed={row['seed']:2d} "
                  f"PPG={row['ppg']:5.1f} E[Pts]={row['expected_pts']:6.1f} "
                  f"Region={row['region']}")
        print()

    # Highlight differences
    ev_players = set(top8_ev['player'])
    div_players = set(top8_div['player'])
    only_ev = ev_players - div_players
    only_div = div_players - ev_players
    if only_ev:
        print(f"  Concentrated adds: {', '.join(only_ev)}")
        print(f"  Diversified swaps in: {', '.join(only_div)}")
    else:
        print("  Both strategies select the same 8 players.")


def run_monte_carlo(df, top_n=8, n_sims=15000):
    """V4: Monte Carlo with team-correlated advancement."""
    print("\n" + "=" * 70)
    print("MONTE CARLO SIMULATION (Team-Correlated, Start-5)")
    print("=" * 70)

    np.random.seed(42)
    top8 = df.head(top_n)

    # Individual player simulations (for per-player stats)
    print(f"\n  Simulating {n_sims:,} tournaments per player...\n")
    for _, row in top8.iterrows():
        sims = simulate_player_enhanced(
            row['ppg'], row['seed'],
            adj_em=row['adj_em'], adj_o=row['adj_o'],
            pace=row['pace'], usage=row['usage'],
            barthag=row['barthag'], three_rate=row['three_rate'],
            n_sims=n_sims)
        print(f"  {row['player']:25s}: Mean={sims.mean():5.1f}, "
              f"Std={sims.std():5.1f}, "
              f"10th-90th=[{np.percentile(sims, 10):.0f} - {np.percentile(sims, 90):.0f}]")

    # V5: Team-correlated lineup simulation
    print(f"\n  --- Lineup Totals (V5: Team-Correlated + Start-5) ---")

    players_data = []
    for _, row in top8.iterrows():
        players_data.append({
            'team': row['team'], 'seed': int(row['seed']),
            'ppg': row['ppg'],
            'adj_em': row['adj_em'] if pd.notna(row['adj_em']) else None,
            'adj_o': row['adj_o'] if pd.notna(row['adj_o']) else None,
            'pace': row['pace'] if pd.notna(row['pace']) else None,
            'usage': row['usage'] if pd.notna(row['usage']) else None,
            'barthag': row['barthag'] if pd.notna(row['barthag']) else None,
            'three_rate': row['three_rate'] if pd.notna(row['three_rate']) else None,
        })

    lineup_totals = simulate_lineup_team_correlated(players_data, n_sims=n_sims)

    # Check team overlap
    team_counts = defaultdict(int)
    for p in players_data:
        team_counts[p['team']] += 1
    stacked = {t: c for t, c in team_counts.items() if c > 1}
    if stacked:
        print(f"  Team stacks detected: {stacked}")
        print(f"  -> These players advance TOGETHER (correlated fates)")

    print(f"  Mean:       {lineup_totals.mean():.1f}")
    print(f"  Median:     {np.median(lineup_totals):.1f}")
    print(f"  Std Dev:    {lineup_totals.std():.1f}")
    print(f"  5th pct:    {np.percentile(lineup_totals, 5):.1f}  (worst case)")
    print(f"  95th pct:   {np.percentile(lineup_totals, 95):.1f}  (best case)")
    print(f"  P(>300):    {np.mean(lineup_totals > 300):.1%}")
    print(f"  P(>350):    {np.mean(lineup_totals > 350):.1%}")
    print(f"  P(>400):    {np.mean(lineup_totals > 400):.1%}")

    return lineup_totals


def run_win_probability_analysis(df, lineup_totals, top_n=8, n_competitors=18):
    """V4: Estimate probability of winning the pool."""
    print("\n" + "=" * 70)
    print("WIN PROBABILITY ANALYSIS (V5)")
    print("=" * 70)

    print(f"\n  Simulating {n_competitors}-team pool competition...")
    print(f"  Opponent strategies: cheat-sheet draft, PPG-greedy, random-smart")
    print(f"  Uses team-correlated advancement for all lineups.\n")

    top8 = df.head(top_n)

    # Build model lineup data
    model_lineup = []
    for _, row in top8.iterrows():
        model_lineup.append({
            'team': row['team'], 'seed': int(row['seed']),
            'ppg': row['ppg'],
            'adj_em': row['adj_em'] if pd.notna(row['adj_em']) else None,
            'adj_o': row['adj_o'] if pd.notna(row['adj_o']) else None,
            'pace': row['pace'] if pd.notna(row['pace']) else None,
            'usage': row['usage'] if pd.notna(row['usage']) else None,
            'barthag': row['barthag'] if pd.notna(row['barthag']) else None,
            'three_rate': row['three_rate'] if pd.notna(row['three_rate']) else None,
        })

    # Build opponent player pool (all players, with scoring params)
    all_player_data = []
    for _, row in df.iterrows():
        all_player_data.append({
            'team': row['team'], 'seed': int(row['seed']),
            'ppg': row['ppg'],
            'adj_em': row['adj_em'] if pd.notna(row['adj_em']) else None,
            'adj_o': row['adj_o'] if pd.notna(row['adj_o']) else None,
            'pace': row['pace'] if pd.notna(row['pace']) else None,
            'usage': row['usage'] if pd.notna(row['usage']) else None,
            'barthag': row['barthag'] if pd.notna(row['barthag']) else None,
            'three_rate': row['three_rate'] if pd.notna(row['three_rate']) else None,
            'rank': int(row['rank']),
        })

    win_prob, model_scores = simulate_win_probability(
        model_lineup, all_player_data, n_competitors=n_competitors, n_sims=2000)

    random_baseline = 1.0 / n_competitors
    edge = win_prob / random_baseline

    print(f"  P(win pool):     {win_prob:.1%}")
    print(f"  Random baseline: {random_baseline:.1%}")
    print(f"  Edge vs random:  {edge:.1f}x")
    print(f"\n  Model lineup scores: Mean={model_scores.mean():.1f}, "
          f"Std={model_scores.std():.1f}")
    print(f"  This means your lineup would be expected to win ~1 in "
          f"{1/max(win_prob, 0.001):.0f} pools.")

    return win_prob


def write_output(df, output_path):
    """Write the ranked draft board to Excel with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Draft Board"

    has_enhanced = df['adj_em'].notna().any() or df['barthag'].notna().any()

    # --- Sheet 1: Draft Board ---
    if has_enhanced:
        headers = ['Draft Rank', 'Player', 'Team', 'Seed', 'Region', 'PPG',
                   'PPG (Shrunk)', 'E[Total Pts]', 'Std Dev', 'E[Games]',
                   'Pts/Game', 'Risk Adj', 'AdjEM', 'Usage%', 'Model',
                   'Shrinkage']
    else:
        headers = ['Draft Rank', 'Player', 'Team', 'Seed', 'Region', 'PPG',
                   'PPG (Shrunk)', 'E[Total Pts]', 'Std Dev', 'E[Games]',
                   'Pts/Game', 'Risk Adj', 'Shrinkage']

    header_font = Font(bold=True, name='Arial', size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    top8_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for r, (_, row) in enumerate(df.iterrows(), 2):
        lam = get_shrinkage_for_seed(row['seed'])
        ppg_shrunk = round((1 - lam) * row['ppg'] + lam * MEAN_PPG, 1)

        if has_enhanced:
            vals = [int(row['draft_rank']), row['player'], row['team'],
                    int(row['seed']), row['region'], round(row['ppg'], 1),
                    ppg_shrunk,
                    round(row['expected_pts'], 1), round(row['std'], 1),
                    round(row['expected_games'], 2), round(row['pts_per_game'], 1),
                    round(row['risk_adj'], 1),
                    round(row['adj_em'], 1) if pd.notna(row['adj_em']) else '',
                    round(row['usage'], 1) if pd.notna(row['usage']) else '',
                    row['model_used'],
                    f"{lam:.0%}"]
        else:
            vals = [int(row['draft_rank']), row['player'], row['team'],
                    int(row['seed']), row['region'], round(row['ppg'], 1),
                    ppg_shrunk,
                    round(row['expected_pts'], 1), round(row['std'], 1),
                    round(row['expected_games'], 2), round(row['pts_per_game'], 1),
                    round(row['risk_adj'], 1),
                    f"{lam:.0%}"]

        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            if c not in (2, 3, 5):
                cell.alignment = Alignment(horizontal='center')
            if r <= 9:
                cell.fill = top8_fill

    # Column widths
    if has_enhanced:
        widths = [10, 28, 22, 6, 10, 6, 10, 14, 10, 10, 10, 10, 8, 8, 10, 10]
    else:
        widths = [10, 28, 22, 6, 10, 6, 10, 14, 10, 10, 10, 10, 10]
    col_letters = [chr(65 + i) for i in range(len(widths))]
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = widths[i]

    ws.freeze_panes = 'A2'

    # --- Sheet 2: Model Info ---
    ws2 = wb.create_sheet("Model Info")
    ws2['A1'] = 'NCAA Tournament Player Pool — V5 Model Parameters'
    ws2['A1'].font = Font(bold=True, size=14, name='Arial')

    info = [
        ('', ''),
        ('Model Version', 'V4 — Bootstrap CV + Team-Correlated Monte Carlo'),
        ('Training Data', '2022-2025 pool results (177 player-season observations)'),
        ('CV Method', 'Bootstrap-Augmented LOYO-CV with 1-SE Rule (200 bootstraps)'),
        ('CV MAE', f'17.00 ± 1.06 (V3 was 17.29)'),
        ('', ''),
        ('PRIMARY MODEL (1-SE Validated)', ''),
        ('Type', f'Ridge regression, alpha=50'),
        ('Features', 'PPG (shrunk) + Seed'),
        ('Alpha (intercept)', f'{ALPHA:.3f}'),
        ('Beta_PPG', f'{BETA_PPG:.3f}'),
        ('Beta_Seed', f'{BETA_SEED:.3f}'),
        ('', ''),
        ('ENHANCED TIEBREAKER MODEL', ''),
        ('Type', f'Ridge regression, alpha=50 (within 1-SE, used when team metrics available)'),
        ('Alpha', f'{ENHANCED_ALPHA:.3f}'),
        ('Beta_PPG', f'{ENHANCED_BETA_PPG:.3f}'),
        ('Beta_Seed', f'{ENHANCED_BETA_SEED:.3f}'),
        ('Beta_Pace', f'{ENHANCED_BETA_PACE:.3f}'),
        ('Beta_Usage', f'{ENHANCED_BETA_USAGE:.3f}'),
        ('Beta_AdjO', f'{ENHANCED_BETA_ADJO:.3f} (near zero — AdjO adds negligible value)'),
        ('', ''),
        ('PER-SEED SHRINKAGE (V4 Improvement)', ''),
        ('Seeds 1-2', f'λ = {SHRINKAGE_TOP:.0%} (reliable PPG, strong schedule)'),
        ('Seeds 3-5', f'λ = {SHRINKAGE_MID:.0%} (less reliable, heavier shrinkage)'),
        ('Seeds 6-8', f'λ = {SHRINKAGE_LOW:.0%}'),
        ('Seeds 9-16', f'λ = {SHRINKAGE_CINDERELLA:.0%} (weak schedule inflates PPG)'),
        ('Population Mean', f'{MEAN_PPG} PPG'),
        ('', ''),
        ('VARIANCE MODEL (HETEROSCEDASTIC)', ''),
        ('Seeds 1-2', f'sigma = {SIGMA_SEEDS_1_2:.2f} pts/game'),
        ('Seeds 3-5', f'sigma = {SIGMA_SEEDS_3_5:.2f} pts/game'),
        ('Seeds 6-8', f'sigma = {SIGMA_SEEDS_6_8:.2f} pts/game'),
        ('Seeds 9-16', f'sigma = {SIGMA_SEEDS_9P:.2f} pts/game'),
        ('', ''),
        ('ADVANCEMENT MODEL (V5)', ''),
        ('Base', 'Historical seed advancement probabilities (1985-2024)'),
        ('BARTHAG (V5 NEW)', f'gamma = {GAMMA_BARTHAG:.3f} — Barttorvik composite win probability'),
        ('Three-Point Rate (V5 NEW)', f'gamma = {GAMMA_3PR:.3f} — only feature with reliable bootstrap CI'),
        ('AdjEM (fallback)', f'gamma = {GAMMA_EM:.3f} — used when BARTHAG unavailable'),
        ('Note', 'BARTHAG subsumes AdjEM (r=0.95). Both validated via 10-method analysis.'),
        ('', ''),
        ('MONTE CARLO (V4/V5)', ''),
        ('Team correlation', 'Players on same team advance TOGETHER (V3 was independent)'),
        ('Start-5 rule', 'Only 5 of 8 players score per round'),
        ('Conditional probs', 'Fixed from V3 (V2 had cumulative probability bug)'),
        ('Win probability', 'Simulates full pool vs opponent draft strategies'),
        ('', ''),
        ('EXPECTED GAMES BY SEED (BASE)', ''),
    ]

    for seed in range(1, 17):
        info.append((f'Seed {seed}', f'{get_base_expected_games(seed):.2f} games'))

    info += [
        ('', ''),
        ('DATA SOURCES FOR ENHANCED COLUMNS', ''),
        ('BARTHAG, 3P Rate', 'barttorvik.com (free) — use pre-tournament values'),
        ('AdjEM, AdjO, Pace', 'kenpom.com ($25/yr) or barttorvik.com (free)'),
        ('Q1 Wins, SOR Rank', 'ncaa.com NCAA NET rankings'),
        ('Usage Rate', 'barttorvik.com or sports-reference.com/cbb (free)'),
    ]

    for r, (label, value) in enumerate(info, 3):
        is_header = any(x in label for x in ['MODEL', 'FORMULA', 'EXPECTED', 'DATA SOURCES',
                                              'ADVANCEMENT', 'VARIANCE', 'MONTE', 'SHRINKAGE',
                                              'TIEBREAKER'])
        ws2.cell(row=r, column=1, value=label).font = Font(
            bold=is_header, name='Arial', size=10)
        ws2.cell(row=r, column=2, value=value).font = Font(name='Arial', size=10)

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 85

    # --- Sheet 3: Team Metrics Reference ---
    if has_enhanced:
        ws3 = wb.create_sheet("Team Metrics")
        ws3['A1'] = 'Team-Level Metrics Used in Draft Board'
        ws3['A1'].font = Font(bold=True, size=14, name='Arial')

        team_headers = ['Team', 'Seed', 'AdjEM', 'AdjO', 'Pace', 'Q1 Wins',
                       'SOR Rank', 'BARTHAG', '3P Rate', 'Adj E[Games]',
                       'Base E[Games]', 'Games Adj']
        for c, h in enumerate(team_headers, 1):
            cell = ws3.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        teams = df.drop_duplicates(subset=['team'])
        # Sort by barthag if available, otherwise adj_em
        sort_col = 'barthag' if df['barthag'].notna().any() else 'adj_em'
        teams = teams.sort_values(sort_col, ascending=False)
        for r, (_, row) in enumerate(teams.iterrows(), 4):
            adj_probs = get_adjusted_advance_probs(
                row['seed'],
                barthag=row['barthag'] if pd.notna(row['barthag']) else None,
                three_rate=row['three_rate'] if pd.notna(row['three_rate']) else None,
                adj_em=row['adj_em'] if pd.notna(row['adj_em']) else None)
            adj_games = sum(adj_probs[:6])
            base_games = get_base_expected_games(row['seed'])

            vals = [row['team'], int(row['seed']),
                    round(row['adj_em'], 1) if pd.notna(row['adj_em']) else '',
                    round(row['adj_o'], 1) if pd.notna(row['adj_o']) else '',
                    round(row['pace'], 1) if pd.notna(row['pace']) else '',
                    int(row['q1_wins']) if pd.notna(row['q1_wins']) else '',
                    int(row['sor_rank']) if pd.notna(row['sor_rank']) else '',
                    round(row['barthag'], 4) if pd.notna(row['barthag']) else '',
                    round(row['three_rate'], 1) if pd.notna(row['three_rate']) else '',
                    round(adj_games, 2), round(base_games, 2),
                    round(adj_games - base_games, 3)]

            for c, val in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=c, value=val)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center')

        widths3 = [22, 6, 8, 8, 8, 10, 10, 10, 10, 14, 14, 12]
        letters3 = [chr(65 + i) for i in range(len(widths3))]
        for i, letter in enumerate(letters3):
            ws3.column_dimensions[letter].width = widths3[i]

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='NCAA Tournament Player Pool — Draft Board Generator (V5)')
    parser.add_argument('input', help='Path to input Excel file (cheat sheet with optional team metrics)')
    parser.add_argument('--output', '-o', default=None,
                        help='Output Excel path (default: draft_board_v5_YYYY.xlsx)')
    parser.add_argument('--risk', '-r', type=float, default=0.3,
                        help='Risk aversion parameter (0=pure EV, 0.3=balanced, 1.0=conservative)')
    args = parser.parse_args()

    if args.output:
        output_path = args.output
    else:
        base = os.path.splitext(os.path.basename(args.input))[0]
        output_path = f'draft_board_v5_{base}.xlsx'

    print("=" * 70)
    print("NCAA TOURNAMENT PLAYER POOL — DRAFT BOARD GENERATOR (V5)")
    print("BARTHAG+3PR Advancement | Team-Correlated MC | Per-Seed Shrinkage")
    print("=" * 70)
    print(f"\n  Input:  {args.input}")
    print(f"  Output: {output_path}")
    print(f"  Risk:   {args.risk}")

    # Load input
    print("\n[1/6] Loading input data...")
    df = load_and_validate_input(args.input)
    print(f"  Loaded {len(df)} players from {args.input}")

    # EDA
    print("\n[2/6] Running exploratory data analysis...")
    run_eda(df)

    # Predictions
    print("\n[3/6] Generating predictions...")
    ranked_df = run_predictions(df)

    # Print top 30 draft board
    print("\n" + "=" * 70)
    print("DRAFT BOARD (Top 30)")
    print("=" * 70)

    has_enhanced = df['adj_em'].notna().any() or df['barthag'].notna().any()
    if has_enhanced:
        print(f"\n  {'Rank':>4} {'Player':25s} {'Team':18s} {'Seed':>4} "
              f"{'PPG':>5} {'E[Pts]':>7} {'Std':>5} {'E[Gm]':>6} "
              f"{'AdjEM':>6} {'λ':>4} {'Region':>8}")
        print("  " + "-" * 108)
        for _, row in ranked_df.head(30).iterrows():
            marker = " <--" if row['draft_rank'] <= 8 else ""
            em_str = f"{row['adj_em']:6.1f}" if pd.notna(row['adj_em']) else "   N/A"
            lam = get_shrinkage_for_seed(row['seed'])
            print(f"  {row['draft_rank']:4.0f} {row['player']:25s} {row['team']:18s} "
                  f"{row['seed']:4d} {row['ppg']:5.1f} {row['expected_pts']:7.1f} "
                  f"{row['std']:5.1f} {row['expected_games']:6.2f} "
                  f"{em_str} {lam:4.0%} {row['region']:>8}{marker}")
    else:
        print(f"\n  {'Rank':>4} {'Player':25s} {'Team':18s} {'Seed':>4} "
              f"{'PPG':>5} {'E[Pts]':>7} {'Std':>5} {'E[Gm]':>8} {'λ':>4} {'Region':>8}")
        print("  " + "-" * 100)
        for _, row in ranked_df.head(30).iterrows():
            marker = " <--" if row['draft_rank'] <= 8 else ""
            lam = get_shrinkage_for_seed(row['seed'])
            print(f"  {row['draft_rank']:4.0f} {row['player']:25s} {row['team']:18s} "
                  f"{row['seed']:4d} {row['ppg']:5.1f} {row['expected_pts']:7.1f} "
                  f"{row['std']:5.1f} {row['expected_games']:8.2f} {lam:4.0%} "
                  f"{row['region']:>8}{marker}")
    print(f"\n  <-- = Top 8 (ideal picks if available)")

    # Optimization
    print("\n[4/6] Running optimization analysis...")
    run_optimization_analysis(ranked_df)

    # Monte Carlo
    print("\n[5/6] Running Monte Carlo simulation (team-correlated)...")
    lineup_totals = run_monte_carlo(ranked_df)

    # Win probability (V4 new)
    print("\n[6/6] Running win probability analysis...")
    win_prob = run_win_probability_analysis(ranked_df, lineup_totals)

    # Write output
    print("\n" + "=" * 70)
    print("WRITING OUTPUT")
    print("=" * 70)
    write_output(ranked_df, output_path)
    print(f"\n  Draft board saved to: {output_path}")
    print(f"  Total players ranked: {len(ranked_df)}")
    if has_enhanced:
        print(f"\n  The Excel contains three sheets:")
        print(f"    1. 'Draft Board'   — Full ranked player list (top 8 highlighted)")
        print(f"    2. 'Model Info'    — V5 model parameters, CV results, shrinkage")
        print(f"    3. 'Team Metrics'  — Team-level AdjEM, Pace, Q1 with adjusted E[Games]")
    else:
        print(f"\n  The Excel contains two sheets:")
        print(f"    1. 'Draft Board' — Full ranked player list (top 8 highlighted)")
        print(f"    2. 'Model Info'  — V5 model parameters and methodology")
    print(f"\n  Use the 'Draft Rank' column as your pick order during the draft.")
    print(f"  When it's your turn, take the highest-ranked available player.")

    print("\n" + "=" * 70)
    print("DONE")
    print("=" * 70)


if __name__ == '__main__':
    main()
