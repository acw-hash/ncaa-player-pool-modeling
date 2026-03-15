#!/usr/bin/env python3
"""
NCAA Tournament Player Pool — Draft Board Generator (V3)
==========================================================
Statistically rigorous draft board generator with cross-validated
Ridge regression, corrected Monte Carlo simulation, and start-5
optimization.

USAGE:
    python draft_board_v2.py input.xlsx
    python draft_board_v2.py input.xlsx --output my_draft_board.xlsx
    python draft_board_v2.py input.xlsx --risk 0.3

INPUT:
    An Excel file with columns (case-insensitive, flexible naming):

    REQUIRED:
        - Player (or Name)
        - Team (or School)
        - Seed
        - PPG (or PTS, Points)
        - Region (optional)
        - Rank (optional)

    OPTIONAL (team-level + player usage — used as tiebreakers):
        - AdjEM (KenPom Adjusted Efficiency Margin, e.g., 28.5)
        - AdjO  (KenPom Adjusted Offensive Efficiency, e.g., 118.5)
        - Pace  (Possessions per 40 min, e.g., 69.8)
        - Q1_Wins (Quadrant 1 wins, e.g., 12)
        - SOR_Rank (Strength of Record ranking, e.g., 1)
        - Usage (or USG, Usage_Pct — player usage rate, e.g., 28.5)

    If optional columns are missing, the model falls back gracefully.
    Fully backward-compatible with V1/V2 input format.

OUTPUT:
    An Excel file with ranked draft board + model diagnostics + team metrics.

MODEL (V3 — Cross-Validated Ridge Regression):
    Per-game scoring: Ridge(PPG + Seed + Pace + Usage + AdjO), alpha=10
    Advancement: Historical seed probs + AdjEM logistic adjustment (gamma=0.08)
    Shrinkage: lambda=0.30 (CV-calibrated; V2 used 0.15)
    Momentum: REMOVED (empirically invalidated — scoring declines in later rounds)
    Monte Carlo: FIXED (uses conditional probabilities, heteroscedastic variance)
    See NCAA_Pool_Enhanced_Analysis_Report_V3.md for full methodology.
"""

import pandas as pd
import numpy as np
import argparse
import sys
import os
from collections import defaultdict

# =============================================================================
# HISTORICAL SEED ADVANCEMENT PROBABILITIES (1985-2024, ~40 years of data)
# Each list: [P(play R64), P(reach R32), P(reach S16), P(reach E8),
#             P(reach F4), P(reach NCG), P(win NCG)]
# =============================================================================
SEED_ADVANCE_PROBS = {
    1:  [1.000, 0.993, 0.850, 0.600, 0.390, 0.220, 0.130],
    2:  [1.000, 0.940, 0.670, 0.400, 0.220, 0.110, 0.060],
    3:  [1.000, 0.850, 0.510, 0.250, 0.120, 0.050, 0.020],
    4:  [1.000, 0.790, 0.430, 0.200, 0.090, 0.040, 0.015],
    5:  [1.000, 0.640, 0.330, 0.140, 0.060, 0.025, 0.010],
    6:  [1.000, 0.630, 0.310, 0.120, 0.050, 0.020, 0.008],
    7:  [1.000, 0.600, 0.230, 0.080, 0.030, 0.012, 0.005],
    8:  [1.000, 0.500, 0.200, 0.070, 0.025, 0.010, 0.004],
    9:  [1.000, 0.500, 0.170, 0.060, 0.020, 0.008, 0.003],
    10: [1.000, 0.390, 0.130, 0.040, 0.015, 0.006, 0.002],
    11: [1.000, 0.370, 0.110, 0.035, 0.012, 0.005, 0.002],
    12: [1.000, 0.360, 0.090, 0.025, 0.008, 0.003, 0.001],
    13: [1.000, 0.210, 0.050, 0.010, 0.003, 0.001, 0.000],
    14: [1.000, 0.150, 0.020, 0.005, 0.001, 0.000, 0.000],
    15: [1.000, 0.070, 0.010, 0.002, 0.000, 0.000, 0.000],
    16: [1.000, 0.010, 0.002, 0.000, 0.000, 0.000, 0.000],
}

# =============================================================================
# PRE-FITTED MODEL PARAMETERS (V3 — Ridge regression, CV-calibrated)
# Trained on 2022-2025, 177 observations, Ridge alpha=10
# All hyperparameters selected via leave-one-year-out cross-validation
# =============================================================================

# --- Per-Game Scoring Model (Ridge Regression, alpha=10) ---
# pts/game = ALPHA + BETA_PPG*ppg_adj + BETA_SEED*seed + BETA_PACE*pace_norm
#           + BETA_USAGE*usage_norm + BETA_ADJO*adjo_norm
# REPRODUCED BY: python v3_model_fitting.py (Step 7)
ALPHA = 0.257           # Intercept
BETA_PPG = 1.007        # Regular-season PPG effect
BETA_SEED = -0.943      # Seed effect on per-game scoring
BETA_PACE = 0.355       # Team pace effect (normalized)
BETA_USAGE = 1.236      # Player usage rate effect (normalized)
BETA_ADJO = -0.117      # Team offensive rating effect (normalized, near zero)

# Heteroscedastic variance by seed bucket (V3 improvement)
SIGMA_SEEDS_1_2 = 3.74  # Per-game scoring std for 1-2 seeds
SIGMA_SEEDS_3_5 = 3.82  # Per-game scoring std for 3-5 seeds
SIGMA_SEEDS_6_8 = 2.85  # Per-game scoring std for 6-8 seeds
SIGMA_SEEDS_9P = 5.93   # Per-game scoring std for 9-16 seeds
SIGMA_GLOBAL = 3.98     # Global RMSE (for fallback)

# Normalization constants (from training data means/stds)
PACE_MEAN = 68.67
PACE_STD = 2.60
USAGE_MEAN = 22.79
USAGE_STD = 4.23
ADJO_MEAN = 115.84
ADJO_STD = 4.45

# --- Bayesian Shrinkage (CV-calibrated) ---
SHRINKAGE = 0.40        # PPG pulled 40% toward population mean (V2 used 0.15)
MEAN_PPG = 14.8         # Population mean PPG
# NOTE: Momentum term REMOVED in V3. Empirical data shows scoring DECLINES
# in later tournament rounds. Setting momentum=0.05 (V2) increased CV MAE by 5.3%.

# --- Advancement Adjustment (AdjEM-based, CV-calibrated) ---
GAMMA_EM = 0.08         # AdjEM effect on log-odds of advancing per round (V2 used 0.026)
ADJEM_MEAN = 21.00      # Training data mean AdjEM
ADJEM_STD = 6.02        # Training data std AdjEM

# --- Fallback: Base Model Parameters (used when team metrics unavailable) ---
# Also Ridge-regularized for consistency
# REPRODUCED BY: python v3_model_fitting.py (Step 7, base model section)
V1_ALPHA = -6.715
V1_BETA_PPG = 1.466
V1_BETA_SEED = -0.888
V1_SIGMA = 4.02


def get_sigma_for_seed(seed):
    """Get the heteroscedastic per-game scoring std for this seed."""
    if seed <= 2:
        return SIGMA_SEEDS_1_2
    elif seed <= 5:
        return SIGMA_SEEDS_3_5
    elif seed <= 8:
        return SIGMA_SEEDS_6_8
    else:
        return SIGMA_SEEDS_9P


def get_base_expected_games(seed):
    """Expected number of tournament games using base seed probabilities."""
    probs = SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16])
    return sum(probs[:6])


def get_adjusted_advance_probs(seed, adj_em=None):
    """
    Get advancement probabilities, optionally adjusted by AdjEM.
    If adj_em is None, returns base seed probabilities.
    """
    probs = list(SEED_ADVANCE_PROBS.get(seed, SEED_ADVANCE_PROBS[16]))

    if adj_em is not None and not np.isnan(adj_em):
        em_norm = (adj_em - ADJEM_MEAN) / ADJEM_STD
        adjusted = [probs[0]]  # R64 always 1.0
        for r in range(1, len(probs)):
            base_p = max(min(probs[r], 0.999), 0.001)
            logit = np.log(base_p / (1 - base_p))
            logit_adj = logit + GAMMA_EM * em_norm
            adj_p = 1 / (1 + np.exp(-logit_adj))
            adj_p = min(adj_p, adjusted[-1])  # Monotonicity
            adjusted.append(adj_p)
        return adjusted
    return probs


def get_game_probabilities(probs):
    """Probability of playing exactly N games (1 through 6) from advancement probs."""
    game_probs = []
    for k in range(6):
        if k < 5:
            game_probs.append(probs[k] - probs[k + 1])
        else:
            game_probs.append(probs[k])
    return game_probs


def predict_player_enhanced(ppg, seed, adj_em=None, adj_o=None, pace=None,
                            usage=None, use_enhanced=True):
    """
    Predict expected tournament total points using the V3 model.

    Returns dict with expected_pts, variance, std, expected_games, risk_adj, etc.
    """
    # Bayesian shrinkage on PPG
    ppg_adj = (1 - SHRINKAGE) * ppg + SHRINKAGE * MEAN_PPG

    # Determine if we have enough team metrics for enhanced model
    has_team_metrics = (use_enhanced and pace is not None and not np.isnan(pace)
                        and usage is not None and not np.isnan(usage))

    if has_team_metrics:
        # Enhanced per-game scoring model (Ridge)
        pace_norm = (pace - PACE_MEAN) / PACE_STD
        usage_norm = (usage - USAGE_MEAN) / USAGE_STD
        adjo_norm = ((adj_o if adj_o is not None and not np.isnan(adj_o) else ADJO_MEAN)
                     - ADJO_MEAN) / ADJO_STD

        pts_per_game = (ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed +
                       BETA_PACE * pace_norm + BETA_USAGE * usage_norm +
                       BETA_ADJO * adjo_norm)
        sigma = get_sigma_for_seed(seed)
        model_used = 'enhanced'
    else:
        # Fallback to base model
        pts_per_game = V1_ALPHA + V1_BETA_PPG * ppg_adj + V1_BETA_SEED * seed
        sigma = get_sigma_for_seed(seed)
        model_used = 'base'

    pts_per_game = max(pts_per_game, 2.0)

    # Adjusted advancement probabilities
    adv_probs = get_adjusted_advance_probs(seed, adj_em)
    exp_games = sum(adv_probs[:6])

    # Expected total points (V3: NO momentum term)
    exp_total = pts_per_game * exp_games + seed

    # Variance (law of total variance)
    game_probs = get_game_probabilities(adv_probs)
    e_games_sq = sum((k + 1) ** 2 * p for k, p in enumerate(game_probs))
    var_games = e_games_sq - exp_games ** 2
    variance = exp_games * sigma ** 2 + var_games * pts_per_game ** 2

    return {
        'expected_pts': round(exp_total, 1),
        'variance': round(variance, 1),
        'std': round(np.sqrt(max(variance, 0)), 1),
        'expected_games': round(exp_games, 2),
        'pts_per_game': round(pts_per_game, 1),
        'risk_adj': round(exp_total - 0.3 * np.sqrt(max(variance, 0)), 1),
        'model_used': model_used,
    }


def simulate_player_enhanced(ppg, seed, adj_em=None, adj_o=None, pace=None,
                              usage=None, n_sims=10000):
    """
    Monte Carlo simulation of a player's tournament total.
    V3 FIX: Uses CONDITIONAL advancement probabilities (not cumulative).
    V3 FIX: Uses heteroscedastic variance by seed bucket.
    V3 FIX: No momentum escalation.
    """
    ppg_adj = (1 - SHRINKAGE) * ppg + SHRINKAGE * MEAN_PPG

    has_team = (pace is not None and not np.isnan(pace)
                and usage is not None and not np.isnan(usage))

    if has_team:
        pace_norm = (pace - PACE_MEAN) / PACE_STD
        usage_norm = (usage - USAGE_MEAN) / USAGE_STD
        adjo_norm = ((adj_o if adj_o is not None and not np.isnan(adj_o) else ADJO_MEAN)
                     - ADJO_MEAN) / ADJO_STD
        pts_per_game_mean = max(ALPHA + BETA_PPG * ppg_adj + BETA_SEED * seed +
                               BETA_PACE * pace_norm + BETA_USAGE * usage_norm +
                               BETA_ADJO * adjo_norm, 2.0)
    else:
        pts_per_game_mean = max(V1_ALPHA + V1_BETA_PPG * ppg_adj + V1_BETA_SEED * seed, 2.0)

    sigma = get_sigma_for_seed(seed)
    probs = get_adjusted_advance_probs(seed, adj_em)

    # V3 FIX: Compute conditional probabilities
    cond_probs = [1.0]  # Always play Round 1
    for rnd in range(1, 6):
        if probs[rnd - 1] > 0:
            cond_probs.append(probs[rnd] / probs[rnd - 1])
        else:
            cond_probs.append(0.0)

    totals = np.zeros(n_sims)
    for sim in range(n_sims):
        total = seed  # seed bonus
        for rnd in range(6):
            # V3 FIX: Use conditional probability for advancement
            if rnd == 0:
                pass  # Always play round 1
            else:
                if np.random.random() > cond_probs[rnd]:
                    break
            # Score this game (V3: no momentum multiplier)
            pts = max(0, np.random.normal(pts_per_game_mean, sigma))
            total += pts
        totals[sim] = total
    return totals


def simulate_lineup_start5(players_data, n_sims=15000):
    """
    Monte Carlo simulation of an 8-player lineup WITH the Start-5 rule.
    Each round, only the top-5 scoring active players count.

    Parameters:
    -----------
    players_data : list of dicts with ppg, seed, adj_em, adj_o, pace, usage

    Returns array of lineup totals across simulations.
    """
    # Precompute per-player parameters
    player_params = []
    for p in players_data:
        ppg_adj = (1 - SHRINKAGE) * p['ppg'] + SHRINKAGE * MEAN_PPG
        pace_val = p.get('pace')
        usage_val = p.get('usage')
        has_team = (pace_val is not None and not np.isnan(pace_val)
                    and usage_val is not None and not np.isnan(usage_val))

        if has_team:
            pace_norm = (pace_val - PACE_MEAN) / PACE_STD
            usage_norm = (usage_val - USAGE_MEAN) / USAGE_STD
            adjo = p.get('adj_o')
            adjo = adjo if adjo is not None and not np.isnan(adjo) else ADJO_MEAN
            adjo_norm = (adjo - ADJO_MEAN) / ADJO_STD
            pts_mean = max(ALPHA + BETA_PPG * ppg_adj + BETA_SEED * p['seed'] +
                          BETA_PACE * pace_norm + BETA_USAGE * usage_norm +
                          BETA_ADJO * adjo_norm, 2.0)
        else:
            pts_mean = max(V1_ALPHA + V1_BETA_PPG * ppg_adj + V1_BETA_SEED * p['seed'], 2.0)

        sigma = get_sigma_for_seed(p['seed'])
        probs = get_adjusted_advance_probs(p['seed'], p.get('adj_em'))

        # Conditional probabilities
        cond_probs = [1.0]
        for rnd in range(1, 6):
            if probs[rnd - 1] > 0:
                cond_probs.append(probs[rnd] / probs[rnd - 1])
            else:
                cond_probs.append(0.0)

        player_params.append({
            'seed': p['seed'], 'pts_mean': pts_mean, 'sigma': sigma,
            'cond_probs': cond_probs,
        })

    n_players = len(player_params)
    lineup_totals = np.zeros(n_sims)

    for sim in range(n_sims):
        total = sum(pp['seed'] for pp in player_params)  # Seed bonuses always count

        # Determine each player's last surviving round via conditional probs
        alive_through = np.zeros(n_players, dtype=int)  # last round index they play
        for i, pp in enumerate(player_params):
            last_rnd = 0  # Always play round 1 (index 0)
            for rnd in range(1, 6):
                if np.random.random() <= pp['cond_probs'][rnd]:
                    last_rnd = rnd
                else:
                    break
            alive_through[i] = last_rnd

        # For each round, score alive players, take top 5
        for rnd in range(6):
            active_scores = []
            for i, pp in enumerate(player_params):
                if alive_through[i] >= rnd:
                    score = max(0, np.random.normal(pp['pts_mean'], pp['sigma']))
                    active_scores.append(score)

            if len(active_scores) == 0:
                break

            # START-5 RULE: Only top 5 count
            active_scores.sort(reverse=True)
            total += sum(active_scores[:5])

        lineup_totals[sim] = total

    return lineup_totals


def load_and_validate_input(filepath):
    """
    Load the input Excel and map columns flexibly.
    Returns a cleaned DataFrame with standardized column names.
    """
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    df = pd.read_excel(filepath)
    df.columns = [str(c).strip() for c in df.columns]

    # Flexible column mapping
    col_map = {}
    for col in df.columns:
        cl = col.lower().replace('_', '').replace(' ', '')
        if cl in ('player', 'name', 'playername'):
            col_map['player'] = col
        elif cl in ('team', 'school'):
            col_map['team'] = col
        elif cl in ('seed',):
            col_map['seed'] = col
        elif cl in ('ppg', 'pts', 'points', 'pts/g', 'avg'):
            col_map['ppg'] = col
        elif cl in ('region', 'reg', 'regions'):
            col_map['region'] = col
        elif cl in ('rank', 'rk', '#'):
            col_map['rank'] = col
        elif cl in ('adjem', 'adjeffmargin', 'kenpomem', 'em'):
            col_map['adj_em'] = col
        elif cl in ('adjo', 'adjoffeff', 'adjoe', 'offeff', 'offrating'):
            col_map['adj_o'] = col
        elif cl in ('pace', 'possessions', 'tempo'):
            col_map['pace'] = col
        elif cl in ('q1wins', 'q1', 'quad1wins', 'quad1'):
            col_map['q1_wins'] = col
        elif cl in ('sorrank', 'sor', 'netrank', 'net', 'strengthofrecord'):
            col_map['sor_rank'] = col
        elif cl in ('usage', 'usg', 'usagepct', 'usg%', 'usagerate'):
            col_map['usage'] = col

    for required in ['player', 'team', 'seed', 'ppg']:
        if required not in col_map:
            print(f"ERROR: Could not find '{required}' column in your Excel.")
            print(f"  Found columns: {list(df.columns)}")
            print(f"  Expected one of: Player/Name, Team/School, Seed, PPG/PTS/Points")
            sys.exit(1)

    rename = {v: k for k, v in col_map.items()}
    df = df.rename(columns=rename)

    df['seed'] = pd.to_numeric(df['seed'], errors='coerce').fillna(16).astype(int)
    df['ppg'] = pd.to_numeric(df['ppg'], errors='coerce').fillna(0.0)
    df['player'] = df['player'].astype(str).str.strip()
    df['team'] = df['team'].astype(str).str.strip()

    if 'region' not in df.columns:
        df['region'] = 'Unknown'
    else:
        df['region'] = df['region'].astype(str).str.strip()

    if 'rank' not in df.columns:
        df['rank'] = range(1, len(df) + 1)

    for opt_col in ['adj_em', 'adj_o', 'pace', 'q1_wins', 'sor_rank', 'usage']:
        if opt_col in df.columns:
            df[opt_col] = pd.to_numeric(df[opt_col], errors='coerce')
        else:
            df[opt_col] = np.nan

    df = df[df['player'].str.len() > 0]
    df = df[df['ppg'] > 0]

    v2_cols_found = [c for c in ['adj_em', 'adj_o', 'pace', 'q1_wins', 'sor_rank', 'usage']
                     if df[c].notna().any()]
    v2_cols_missing = [c for c in ['adj_em', 'adj_o', 'pace', 'q1_wins', 'sor_rank', 'usage']
                       if not df[c].notna().any()]

    print(f"\n  --- Column Detection ---")
    print(f"  Required columns: Player, Team, Seed, PPG [OK]")
    if v2_cols_found:
        print(f"  Enhanced columns found: {', '.join(v2_cols_found)}")
        coverage = {c: f"{df[c].notna().mean():.0%}" for c in v2_cols_found}
        print(f"  Coverage: {coverage}")
    if v2_cols_missing:
        print(f"  Enhanced columns missing (using defaults): {', '.join(v2_cols_missing)}")

    std_cols = ['player', 'team', 'seed', 'region', 'ppg', 'rank',
                'adj_em', 'adj_o', 'pace', 'q1_wins', 'sor_rank', 'usage']
    return df[std_cols].reset_index(drop=True)


def run_eda(df):
    """Print exploratory data analysis to console."""
    print("\n" + "=" * 70)
    print("EXPLORATORY DATA ANALYSIS")
    print("=" * 70)

    print(f"\nTotal players loaded: {len(df)}")
    print(f"Seeds represented: {sorted(df['seed'].unique())}")
    print(f"Regions: {sorted(df['region'].unique())}")

    print("\n--- Player Count by Seed ---")
    seed_counts = df.groupby('seed').size()
    for seed, count in seed_counts.items():
        exp_g = get_base_expected_games(seed)
        sigma = get_sigma_for_seed(seed)
        print(f"  Seed {seed:2d}: {count:3d} players | E[games]: {exp_g:.2f} | "
              f"Per-game sigma: {sigma:.2f}")

    print("\n--- PPG Distribution by Seed Bucket ---")
    buckets = {
        'Top (1-2)': df[df['seed'].isin([1, 2])],
        'Mid (3-5)': df[df['seed'].isin([3, 4, 5])],
        'Low (6-8)': df[df['seed'].isin([6, 7, 8])],
        'Cinderella (9+)': df[df['seed'] >= 9],
    }
    for name, bucket in buckets.items():
        if len(bucket) > 0:
            print(f"  {name:20s}: N={len(bucket):3d}, "
                  f"PPG mean={bucket['ppg'].mean():.1f}, "
                  f"PPG range=[{bucket['ppg'].min():.1f} - {bucket['ppg'].max():.1f}]")

    has_em = df['adj_em'].notna().any()
    if has_em:
        print("\n--- Team Metrics Summary ---")
        for col, label in [('adj_em', 'AdjEM'), ('adj_o', 'AdjO'), ('pace', 'Pace'),
                          ('q1_wins', 'Q1 Wins'), ('usage', 'Usage%')]:
            if df[col].notna().any():
                vals = df[col].dropna()
                print(f"  {label:10s}: mean={vals.mean():.1f}, "
                      f"range=[{vals.min():.1f} - {vals.max():.1f}], "
                      f"coverage={vals.notna().sum()}/{len(df)}")

    if 'region' in df.columns:
        print("\n--- Players by Region ---")
        for region, group in df.groupby('region'):
            seeds = group['seed'].value_counts().sort_index()
            seed_str = ', '.join(f"{s}-seed: {c}" for s, c in seeds.items() if s <= 4)
            em_str = ""
            if has_em and group['adj_em'].notna().any():
                em_str = f" | Avg AdjEM: {group['adj_em'].mean():.1f}"
            print(f"  {region:8s}: {len(group):3d} players | {seed_str}{em_str}")


def run_predictions(df):
    """Generate predictions for all players. Returns df with prediction columns."""
    print("\n" + "=" * 70)
    print("PREDICTIVE MODEL (V3 — CROSS-VALIDATED RIDGE REGRESSION)")
    print("=" * 70)

    has_enhanced = df['pace'].notna().any() and df['usage'].notna().any()

    if has_enhanced:
        print("\n  Enhanced model ACTIVE — using team metrics + usage rate as tiebreakers")
        print(f"\n  Per-game scoring: pts/game = {ALPHA:.3f} + {BETA_PPG:.3f}*PPG_adj "
              f"+ {BETA_SEED:.3f}*Seed")
        print(f"    + {BETA_PACE:.3f}*Pace_norm + {BETA_USAGE:.3f}*Usage_norm "
              f"+ {BETA_ADJO:.3f}*AdjO_norm")
        print(f"  Advancement: Seed lookup + AdjEM adjustment (gamma={GAMMA_EM:.3f})")
    else:
        print("\n  Base model — enhanced columns not found (using PPG + Seed)")
        print(f"  Per-game: pts/game = {V1_ALPHA:.3f} + {V1_BETA_PPG:.3f}*PPG_adj "
              f"+ {V1_BETA_SEED:.3f}*Seed")

    print(f"\n  Bayesian shrinkage: {SHRINKAGE:.0%} toward mean PPG of {MEAN_PPG:.1f}")
    print(f"  Momentum: NONE (empirically invalidated — scoring declines in later rounds)")
    print(f"  Variance: Heteroscedastic by seed bucket (sigma 1-2={SIGMA_SEEDS_1_2:.2f}, "
          f"3-5={SIGMA_SEEDS_3_5:.2f}, 6-8={SIGMA_SEEDS_6_8:.2f}, 9+={SIGMA_SEEDS_9P:.2f})")

    # Demo prediction
    print(f"\n  --- Demo: 18 PPG player on a 2-seed ---")
    demo_args = {'ppg': 18.0, 'seed': 2}
    if has_enhanced:
        demo_args.update({'adj_em': 22.0, 'adj_o': 115.0, 'pace': 68.0, 'usage': 24.0})
    demo = predict_player_enhanced(**demo_args)
    print(f"    -> {demo['pts_per_game']} projected pts/game")
    print(f"    -> {demo['expected_games']} expected games")
    print(f"    -> {demo['expected_pts']} expected total points")
    print(f"    -> Model used: {demo['model_used']}")

    # Generate predictions for all players
    predictions = []
    enhanced_count = 0
    base_count = 0
    for _, row in df.iterrows():
        pred = predict_player_enhanced(
            ppg=row['ppg'],
            seed=row['seed'],
            adj_em=row['adj_em'] if pd.notna(row['adj_em']) else None,
            adj_o=row['adj_o'] if pd.notna(row['adj_o']) else None,
            pace=row['pace'] if pd.notna(row['pace']) else None,
            usage=row['usage'] if pd.notna(row['usage']) else None,
        )
        predictions.append(pred)
        if pred['model_used'] == 'enhanced':
            enhanced_count += 1
        else:
            base_count += 1

    print(f"\n  Players scored with enhanced model: {enhanced_count}")
    print(f"  Players scored with base model:     {base_count}")

    pred_df = pd.DataFrame(predictions)
    result = pd.concat([df.reset_index(drop=True), pred_df], axis=1)
    result = result.sort_values('expected_pts', ascending=False).reset_index(drop=True)
    result['draft_rank'] = range(1, len(result) + 1)

    return result


def run_optimization_analysis(df):
    """Compare optimization strategies including Start-5 rule."""
    print("\n" + "=" * 70)
    print("OPTIMIZATION STRATEGY COMPARISON")
    print("=" * 70)

    print("\n  Selecting top 8 under different strategies.\n")

    # Strategy 1: Pure expected value
    top8_ev = df.head(8).copy()

    # Strategy 2: Diversified (max 3 per region, max 3 per seed)
    selected = []
    region_counts = defaultdict(int)
    seed_counts = defaultdict(int)
    for _, row in df.iterrows():
        if len(selected) >= 8:
            break
        if region_counts[row['region']] >= 3:
            continue
        if seed_counts[row['seed']] >= 3:
            continue
        selected.append(row)
        region_counts[row['region']] += 1
        seed_counts[row['seed']] += 1
    top8_div = pd.DataFrame(selected)

    # Strategy 3: AdjEM-weighted
    if df['adj_em'].notna().any():
        df_em = df.copy()
        em_rank = df_em['adj_em'].rank(ascending=False, na_option='bottom')
        df_em['em_boosted'] = df_em['expected_pts'] + (len(df_em) - em_rank) * 0.05
        top8_em = df_em.nlargest(8, 'em_boosted')
    else:
        top8_em = None

    strategies = {
        'High Upside (Concentrated — RECOMMENDED)': top8_ev,
        'Balanced (Diversified across regions/seeds)': top8_div,
    }
    if top8_em is not None:
        strategies['AdjEM-Weighted (Team Strength Bias)'] = top8_em

    for name, lineup in strategies.items():
        total_exp = lineup['expected_pts'].sum()
        total_std = np.sqrt(lineup['variance'].sum())
        print(f"  === {name} ===")
        print(f"  Total E[Pts]: {total_exp:.1f} | Total Std: {total_std:.1f}\n")
        for _, row in lineup.iterrows():
            em_str = f"AdjEM={row['adj_em']:.0f}" if pd.notna(row['adj_em']) else ""
            usg_str = f"USG={row['usage']:.0f}%" if pd.notna(row['usage']) else ""
            extra = f" | {em_str} {usg_str}".strip(' |') if (em_str or usg_str) else ""
            print(f"    {row['draft_rank']:3.0f}. {row['player']:25s} "
                  f"{row['team']:18s} Seed={row['seed']:2d} "
                  f"PPG={row['ppg']:5.1f} E[Pts]={row['expected_pts']:6.1f} "
                  f"Region={row['region']}{extra}")
        print()


def run_monte_carlo(df, top_n=8, n_sims=15000):
    """Monte Carlo simulation with corrected conditional probabilities and Start-5 rule."""
    print("\n" + "=" * 70)
    print("MONTE CARLO SIMULATION (Top 8 Lineup)")
    print("=" * 70)

    np.random.seed(42)
    top8 = df.head(top_n)

    # First: standard per-player simulation (no start-5 constraint)
    print(f"\n  Simulating {n_sims:,} tournaments (corrected conditional probabilities)...\n")
    lineup_totals = np.zeros(n_sims)

    for _, row in top8.iterrows():
        sims = simulate_player_enhanced(
            ppg=row['ppg'], seed=row['seed'],
            adj_em=row['adj_em'] if pd.notna(row['adj_em']) else None,
            adj_o=row['adj_o'] if pd.notna(row['adj_o']) else None,
            pace=row['pace'] if pd.notna(row['pace']) else None,
            usage=row['usage'] if pd.notna(row['usage']) else None,
            n_sims=n_sims,
        )
        lineup_totals += sims

        print(f"  {row['player']:25s}: Mean={sims.mean():5.1f}, "
              f"Std={sims.std():5.1f}, "
              f"10th-90th=[{np.percentile(sims, 10):.0f} - {np.percentile(sims, 90):.0f}]")

    print(f"\n  --- Lineup Totals (All 8 Scoring — Upper Bound) ---")
    print(f"  Mean:       {lineup_totals.mean():.1f}")
    print(f"  Median:     {np.median(lineup_totals):.1f}")
    print(f"  Std Dev:    {lineup_totals.std():.1f}")
    print(f"  5th pct:    {np.percentile(lineup_totals, 5):.1f}  (worst case)")
    print(f"  95th pct:   {np.percentile(lineup_totals, 95):.1f}  (best case)")
    print(f"  P(>300):    {np.mean(lineup_totals > 300):.1%}")
    print(f"  P(>350):    {np.mean(lineup_totals > 350):.1%}")
    print(f"  P(>400):    {np.mean(lineup_totals > 400):.1%}")

    # Second: Start-5 simulation
    print(f"\n  --- Lineup Totals (Start-5 Rule — Realistic) ---")
    players_data = []
    for _, row in top8.iterrows():
        players_data.append({
            'ppg': row['ppg'], 'seed': row['seed'],
            'adj_em': row['adj_em'] if pd.notna(row['adj_em']) else None,
            'adj_o': row['adj_o'] if pd.notna(row['adj_o']) else None,
            'pace': row['pace'] if pd.notna(row['pace']) else None,
            'usage': row['usage'] if pd.notna(row['usage']) else None,
        })
    start5_totals = simulate_lineup_start5(players_data, n_sims=n_sims)
    print(f"  Mean:       {start5_totals.mean():.1f}")
    print(f"  Median:     {np.median(start5_totals):.1f}")
    print(f"  Std Dev:    {start5_totals.std():.1f}")
    print(f"  5th pct:    {np.percentile(start5_totals, 5):.1f}  (worst case)")
    print(f"  95th pct:   {np.percentile(start5_totals, 95):.1f}  (best case)")


def write_output(df, output_path):
    """Write the ranked draft board to Excel with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Draft Board ---
    ws = wb.active
    ws.title = "Draft Board"

    has_enhanced = df['adj_em'].notna().any()

    headers = ['Draft Rank', 'Player', 'Team', 'Seed', 'Region', 'Reg Season PPG',
               'Expected Total Pts', 'Std Dev', 'Expected Games', 'Proj Pts/Game',
               'Risk-Adjusted Score', 'Cheat Sheet Rank']
    cols = ['draft_rank', 'player', 'team', 'seed', 'region', 'ppg',
            'expected_pts', 'std', 'expected_games', 'pts_per_game',
            'risk_adj', 'rank']

    if has_enhanced:
        headers += ['AdjEM', 'AdjO', 'Pace', 'Q1 Wins', 'SOR Rank', 'Usage%', 'Model']
        cols += ['adj_em', 'adj_o', 'pace', 'q1_wins', 'sor_rank', 'usage', 'model_used']

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    top8_fill = PatternFill('solid', fgColor='D6E4F0')
    data_font = Font(name='Arial', size=10)
    border = Border(
        bottom=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
    )

    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            val = row.get(col, '')
            if isinstance(val, (np.integer, np.int64)):
                val = int(val)
            elif isinstance(val, (np.floating, np.float64)):
                if np.isnan(val):
                    val = ''
                else:
                    val = float(val)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c in (1, 4, 7, 8, 9, 10, 11, 12):
                cell.alignment = Alignment(horizontal='center')
            if r <= 9:
                cell.fill = top8_fill

    base_widths = [10, 28, 22, 6, 10, 14, 16, 10, 14, 14, 16, 16]
    if has_enhanced:
        base_widths += [8, 8, 8, 8, 10, 10, 10]

    col_letters = [chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
                   for i in range(len(base_widths))]
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = base_widths[i]

    ws.freeze_panes = 'A2'

    # --- Sheet 2: Model Info ---
    ws2 = wb.create_sheet("Model Info")
    ws2['A1'] = 'NCAA Tournament Player Pool — V3 Model Parameters (Cross-Validated Ridge)'
    ws2['A1'].font = Font(bold=True, size=14, name='Arial')

    info = [
        ('', ''),
        ('Model Version', 'V3 — Cross-Validated Ridge Regression (alpha=10)'),
        ('Training Data', '2022-2025 pool results (177 player-season observations)'),
        ('CV Performance', 'LOYO CV MAE=17.29, CV r=0.41 (avg across 4 held-out years)'),
        ('Key Changes from V2', 'Momentum REMOVED, shrinkage increased to 30%, Ridge regression, fixed MC'),
        ('', ''),
        ('PER-GAME SCORING MODEL (RIDGE)', ''),
        ('Formula', f'pts/game = {ALPHA:.3f} + {BETA_PPG:.3f}*PPG_adj + {BETA_SEED:.3f}*Seed + {BETA_PACE:.3f}*Pace_norm + {BETA_USAGE:.3f}*Usage_norm + {BETA_ADJO:.3f}*AdjO_norm'),
        ('Alpha (intercept)', f'{ALPHA:.3f}'),
        ('Beta_PPG', f'{BETA_PPG:.3f} — each regular-season PPG contributes this many tournament pts/game'),
        ('Beta_Seed', f'{BETA_SEED:.3f} — higher seeds score less per game'),
        ('Beta_Pace', f'{BETA_PACE:.3f} — team pace effect (normalized, mean={PACE_MEAN:.1f}, std={PACE_STD:.1f})'),
        ('Beta_Usage', f'{BETA_USAGE:.3f} — player usage rate effect (normalized, mean={USAGE_MEAN:.1f}, std={USAGE_STD:.1f})'),
        ('Beta_AdjO', f'{BETA_ADJO:.3f} — team offensive rating effect (normalized, mean={ADJO_MEAN:.1f}, std={ADJO_STD:.1f})'),
        ('Ridge Alpha', '10.0 — regularization prevents overfitting with correlated features'),
        ('Shrinkage', f'{SHRINKAGE:.2f} — PPG pulled {SHRINKAGE:.0%} toward mean of {MEAN_PPG:.1f} (CV-calibrated)'),
        ('Momentum', '0.00 — REMOVED. Empirical data shows scoring DECLINES in later rounds.'),
        ('', ''),
        ('VARIANCE MODEL (HETEROSCEDASTIC)', ''),
        ('Seeds 1-2', f'sigma = {SIGMA_SEEDS_1_2:.2f} pts/game'),
        ('Seeds 3-5', f'sigma = {SIGMA_SEEDS_3_5:.2f} pts/game'),
        ('Seeds 6-8', f'sigma = {SIGMA_SEEDS_6_8:.2f} pts/game'),
        ('Seeds 9-16', f'sigma = {SIGMA_SEEDS_9P:.2f} pts/game (58% higher than 1-2 seeds)'),
        ('', ''),
        ('ADVANCEMENT MODEL', ''),
        ('Base', 'Historical seed advancement probabilities (1985-2024, 40 years)'),
        ('Enhancement', f'AdjEM logistic adjustment: gamma_em = {GAMMA_EM:.3f} (CV-calibrated)'),
        ('AdjEM Normalization', f'mean = {ADJEM_MEAN:.2f}, std = {ADJEM_STD:.2f}'),
        ('', ''),
        ('MONTE CARLO', ''),
        ('V3 Fix', 'Uses CONDITIONAL probabilities (V2 used cumulative — was wrong)'),
        ('V3 Fix', 'Heteroscedastic variance by seed bucket'),
        ('V3 Addition', 'Start-5 simulation (only 5 of 8 players score per round)'),
        ('', ''),
        ('EXPECTED GAMES BY SEED (BASE)', ''),
    ]

    for seed in range(1, 17):
        info.append((f'Seed {seed}', f'{get_base_expected_games(seed):.2f} games'))

    info += [
        ('', ''),
        ('DATA SOURCES FOR ENHANCED COLUMNS', ''),
        ('AdjEM, AdjO, Pace', 'kenpom.com ($25/yr) or barttorvik.com (free)'),
        ('Q1 Wins, SOR Rank', 'ncaa.com/rankings/basketball-men/d1/ncaa-mens-basketball-net-rankings'),
        ('Usage Rate', 'barttorvik.com or sports-reference.com/cbb (free)'),
    ]

    for r, (label, value) in enumerate(info, 3):
        is_header = any(x in label for x in ['MODEL', 'FORMULA', 'EXPECTED', 'DATA SOURCES',
                                              'ADVANCEMENT', 'VARIANCE', 'MONTE'])
        ws2.cell(row=r, column=1, value=label).font = Font(
            bold=is_header, name='Arial', size=10
        )
        ws2.cell(row=r, column=2, value=value).font = Font(name='Arial', size=10)

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 85

    # --- Sheet 3: Team Metrics Reference ---
    if has_enhanced:
        ws3 = wb.create_sheet("Team Metrics")
        ws3['A1'] = 'Team-Level Metrics Used in Draft Board'
        ws3['A1'].font = Font(bold=True, size=14, name='Arial')

        team_headers = ['Team', 'Seed', 'AdjEM', 'AdjO', 'Pace', 'Q1 Wins',
                       'SOR Rank', 'Adj E[Games]', 'Base E[Games]', 'Games Adjustment']
        for c, h in enumerate(team_headers, 1):
            cell = ws3.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        teams = df.drop_duplicates(subset=['team']).sort_values('adj_em', ascending=False)
        for r, (_, row) in enumerate(teams.iterrows(), 4):
            adj_probs = get_adjusted_advance_probs(
                row['seed'],
                row['adj_em'] if pd.notna(row['adj_em']) else None
            )
            adj_games = sum(adj_probs[:6])
            base_games = get_base_expected_games(row['seed'])

            vals = [row['team'], int(row['seed']),
                    round(row['adj_em'], 1) if pd.notna(row['adj_em']) else '',
                    round(row['adj_o'], 1) if pd.notna(row['adj_o']) else '',
                    round(row['pace'], 1) if pd.notna(row['pace']) else '',
                    int(row['q1_wins']) if pd.notna(row['q1_wins']) else '',
                    int(row['sor_rank']) if pd.notna(row['sor_rank']) else '',
                    round(adj_games, 2), round(base_games, 2),
                    round(adj_games - base_games, 3)]

            for c, val in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=c, value=val)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center')

        widths3 = [22, 6, 8, 8, 8, 10, 10, 14, 14, 16]
        letters3 = [chr(65 + i) for i in range(len(widths3))]
        for i, letter in enumerate(letters3):
            ws3.column_dimensions[letter].width = widths3[i]

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='NCAA Tournament Player Pool — Draft Board Generator (V3)')
    parser.add_argument('input', help='Path to input Excel file (cheat sheet with optional team metrics)')
    parser.add_argument('--output', '-o', default=None,
                        help='Output Excel path (default: draft_board_v3_YYYY.xlsx)')
    parser.add_argument('--risk', '-r', type=float, default=0.3,
                        help='Risk aversion parameter (0=pure EV, 0.3=balanced, 1.0=conservative)')
    args = parser.parse_args()

    if args.output:
        output_path = args.output
    else:
        base = os.path.splitext(os.path.basename(args.input))[0]
        output_path = f'draft_board_v3_{base}.xlsx'

    print("=" * 70)
    print("NCAA TOURNAMENT PLAYER POOL — DRAFT BOARD GENERATOR (V3)")
    print("Cross-Validated Ridge Regression | Corrected Monte Carlo")
    print("=" * 70)
    print(f"\n  Input:  {args.input}")
    print(f"  Output: {output_path}")
    print(f"  Risk:   {args.risk}")

    # Load input
    print("\n[1/5] Loading input data...")
    df = load_and_validate_input(args.input)
    print(f"  Loaded {len(df)} players from {args.input}")

    # EDA
    print("\n[2/5] Running exploratory data analysis...")
    run_eda(df)

    # Predictions
    print("\n[3/5] Generating predictions...")
    ranked_df = run_predictions(df)

    # Print top 30 draft board
    print("\n" + "=" * 70)
    print("DRAFT BOARD (Top 30)")
    print("=" * 70)

    has_enhanced = df['adj_em'].notna().any()
    if has_enhanced:
        print(f"\n  {'Rank':>4} {'Player':25s} {'Team':18s} {'Seed':>4} "
              f"{'PPG':>5} {'E[Pts]':>7} {'Std':>5} {'E[Gm]':>6} "
              f"{'AdjEM':>6} {'USG%':>5} {'Region':>8}")
        print("  " + "-" * 110)
        for _, row in ranked_df.head(30).iterrows():
            marker = " <--" if row['draft_rank'] <= 8 else ""
            em_str = f"{row['adj_em']:6.1f}" if pd.notna(row['adj_em']) else "   N/A"
            usg_str = f"{row['usage']:5.1f}" if pd.notna(row['usage']) else "  N/A"
            print(f"  {row['draft_rank']:4.0f} {row['player']:25s} {row['team']:18s} "
                  f"{row['seed']:4d} {row['ppg']:5.1f} {row['expected_pts']:7.1f} "
                  f"{row['std']:5.1f} {row['expected_games']:6.2f} "
                  f"{em_str} {usg_str} {row['region']:>8}{marker}")
    else:
        print(f"\n  {'Rank':>4} {'Player':25s} {'Team':18s} {'Seed':>4} "
              f"{'PPG':>5} {'E[Pts]':>7} {'Std':>5} {'E[Games]':>8} {'Region':>8}")
        print("  " + "-" * 95)
        for _, row in ranked_df.head(30).iterrows():
            marker = " <--" if row['draft_rank'] <= 8 else ""
            print(f"  {row['draft_rank']:4.0f} {row['player']:25s} {row['team']:18s} "
                  f"{row['seed']:4d} {row['ppg']:5.1f} {row['expected_pts']:7.1f} "
                  f"{row['std']:5.1f} {row['expected_games']:8.2f} {row['region']:>8}{marker}")
    print(f"\n  <-- = Top 8 (ideal picks if available)")

    # Optimization
    print("\n[4/5] Running optimization analysis...")
    run_optimization_analysis(ranked_df)

    # Monte Carlo
    print("\n[5/5] Running Monte Carlo simulation...")
    run_monte_carlo(ranked_df)

    # Write output
    print("\n" + "=" * 70)
    print("WRITING OUTPUT")
    print("=" * 70)
    write_output(ranked_df, output_path)
    print(f"\n  Draft board saved to: {output_path}")
    print(f"  Total players ranked: {len(ranked_df)}")
    if has_enhanced:
        print(f"\n  The Excel contains three sheets:")
        print(f"    1. 'Draft Board'   — Full ranked player list (top 8 highlighted)")
        print(f"    2. 'Model Info'    — V3 model parameters and CV results")
        print(f"    3. 'Team Metrics'  — Team-level AdjEM, Pace, Q1 with adjusted E[Games]")
    else:
        print(f"\n  The Excel contains two sheets:")
        print(f"    1. 'Draft Board' — Full ranked player list (top 8 highlighted)")
        print(f"    2. 'Model Info'  — V3 model parameters and methodology")
    print(f"\n  Use the 'Draft Rank' column as your pick order during the draft.")
    print(f"  When it's your turn, take the highest-ranked available player.")

    print("\n" + "=" * 70)
    print("DONE")
    print("=" * 70)


if __name__ == '__main__':
    main()
